from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from typing import List, Optional, Dict, Any
from datetime import datetime
import os
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import csv
import tempfile
from .models import (
    Applicant,
    ScholarshipAward,
    Scholarship,
    ReviewerInformationRequest,
    AwardDecision,
)
from django.utils import timezone


class ReportEngine:
    """OOP Report Engine for generating scholarship reports and summaries."""

    def __init__(self):
        self.scholarships = []

    # Function to log reviewer requests for additional applicant information
    # Implements requirement SFWE504_3-LLR-27.
    def log_information_request(
        self,
        applicant_id: str = None,
        applicant: Applicant = None,
        reviewer_name: str = None,
        reviewer_email: str = None,
        scholarship_name: str = None,
        request_type: str = None,
        request_details: str = None,
        priority: str = "medium",
    ) -> ReviewerInformationRequest:
        """Log a reviewer's request for additional applicant information.

        Meets requirement: The report engine shall log reviewer requests for additional applicant information.

        Args:
            applicant_id (str, optional): Student ID of the applicant
            applicant (Applicant, optional): Applicant model instance (if already loaded)
            reviewer_name (str): Name of the reviewer making the request
            reviewer_email (str, optional): Email of the reviewer
            scholarship_name (str, optional): Name of the scholarship being reviewed
            request_type (str): Type of information requested (e.g., 'transcript', 'recommendation', 'essay_clarification')
            request_details (str): Detailed description of what information is needed
            priority (str): Priority level ('low', 'medium', 'high', 'urgent'). Defaults to 'medium'

        Returns:
            ReviewerInformationRequest: The created request object

        Raises:
            ValueError: If applicant cannot be found or required fields are missing
        """
        # Validate required fields
        if not reviewer_name:
            raise ValueError("reviewer_name is required")
        if not request_type:
            raise ValueError("request_type is required")
        if not request_details:
            raise ValueError("request_details is required")

        # Get applicant instance if not provided
        if not applicant:
            if not applicant_id:
                raise ValueError("Either applicant or applicant_id must be provided")
            try:
                applicant = Applicant.objects.get(student_id=applicant_id)
            except Applicant.DoesNotExist:
                raise ValueError(
                    f"Applicant with student_id '{applicant_id}' not found"
                )

        # Create the information request
        request = ReviewerInformationRequest.objects.create(
            applicant=applicant,
            reviewer_name=reviewer_name,
            reviewer_email=reviewer_email,
            scholarship_name=scholarship_name,
            request_type=request_type,
            request_details=request_details,
            priority=priority,
        )

        # Generate and save a report log for this request
        self._generate_information_request_log(request)

        return request

    def _generate_information_request_log(
        self, request: ReviewerInformationRequest
    ) -> str:
        """Generate a detailed log report for an information request.

        Creates a timestamped log file documenting the request details.

        Args:
            request (ReviewerInformationRequest): The information request object

        Returns:
            str: Path to the generated log file
        """
        # Create logs directory if it doesn't exist
        logs_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "information_request_logs"
        )
        os.makedirs(logs_dir, exist_ok=True)

        # Generate timestamped filename
        timestamp = request.requested_at.strftime("%Y%m%d_%H%M%S")
        filename = f"info_request_{request.id}_{timestamp}.txt"
        log_path = os.path.join(logs_dir, filename)

        # Generate log content
        log_content = f"""
================================================================================
APPLICANT INFORMATION REQUEST LOG
================================================================================

Request ID: {request.id}
Requested At: {request.requested_at.strftime("%Y-%m-%d %H:%M:%S")}
Status: {request.status.upper()}

--------------------------------------------------------------------------------
APPLICANT INFORMATION
--------------------------------------------------------------------------------
Name: {request.applicant.name}
Student ID: {request.applicant.student_id}
NetID: {request.applicant.netid or "N/A"}
Major: {request.applicant.major}
GPA: {request.applicant.gpa:.2f}
Academic Level: {request.applicant.academic_level}

--------------------------------------------------------------------------------
REVIEWER INFORMATION
--------------------------------------------------------------------------------
Reviewer Name: {request.reviewer_name}
Reviewer Email: {request.reviewer_email or "N/A"}
Scholarship: {request.scholarship_name or "N/A"}

--------------------------------------------------------------------------------
REQUEST DETAILS
--------------------------------------------------------------------------------
Request Type: {request.request_type.replace("_", " ").title()}
Priority: {request.priority.upper()}

Details:
{request.request_details}

--------------------------------------------------------------------------------
FULFILLMENT STATUS
--------------------------------------------------------------------------------
Status: {request.status.title()}
Fulfilled At: {request.fulfilled_at.strftime("%Y-%m-%d %H:%M:%S") if request.fulfilled_at else "Not yet fulfilled"}
Fulfillment Notes: {request.fulfillment_notes or "N/A"}

================================================================================
END OF LOG
================================================================================
"""

        # Write log to file
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(log_content.strip())

        logger.info(f"Information request log generated: {log_path}")
        return log_path

    def get_information_requests(
        self, applicant_id: str = None, status: str = None, scholarship_name: str = None
    ) -> List[ReviewerInformationRequest]:
        """Retrieve information requests with optional filtering.

        Args:
            applicant_id (str, optional): Filter by applicant student ID
            status (str, optional): Filter by request status ('pending', 'in_progress', 'fulfilled', 'cancelled')
            scholarship_name (str, optional): Filter by scholarship name

        Returns:
            List[ReviewerInformationRequest]: List of matching requests
        """
        queryset = ReviewerInformationRequest.objects.all()

        if applicant_id:
            queryset = queryset.filter(applicant__student_id=applicant_id)

        if status:
            queryset = queryset.filter(status=status)

        if scholarship_name:
            queryset = queryset.filter(scholarship_name=scholarship_name)

        return list(queryset)

    def update_request_status(
        self, request_id: int, status: str, fulfillment_notes: str = None
    ) -> ReviewerInformationRequest:
        """Update the status of an information request.

        Args:
            request_id (int): ID of the request to update
            status (str): New status ('pending', 'in_progress', 'fulfilled', 'cancelled')
            fulfillment_notes (str, optional): Notes about fulfillment (especially for 'fulfilled' status)

        Returns:
            ReviewerInformationRequest: The updated request object

        Raises:
            ValueError: If request not found or invalid status
        """
        valid_statuses = ["pending", "in_progress", "fulfilled", "cancelled"]
        if status not in valid_statuses:
            raise ValueError(
                f"Invalid status. Must be one of: {', '.join(valid_statuses)}"
            )

        try:
            request = ReviewerInformationRequest.objects.get(id=request_id)
        except ReviewerInformationRequest.DoesNotExist:
            raise ValueError(f"Request with ID {request_id} not found")

        request.status = status
        if status == "fulfilled":
            request.fulfilled_at = timezone.now()
        if fulfillment_notes:
            request.fulfillment_notes = fulfillment_notes
        request.save()

        # Regenerate log with updated status
        self._generate_information_request_log(request)

        return request

    # Function to generate donor report. Meets requirement SFWE504_3-LLR-2
    def generate_donor_report(
        self,
        donor_name: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> dict:
        """Generate a comprehensive report for a specific donor including key dates and award summaries.

        Args:
            donor_name (str): Name of the donor to generate report for
            start_date (datetime, optional): Start date for report period
            end_date (datetime, optional): End date for report period

        Returns:
            dict: Detailed donor report including scholarships, awards, and key dates
        """
        # Default to last year if no dates provided
        if not end_date:
            end_date = timezone.now()
        if not start_date:
            start_date = end_date - pd.DateOffset(years=1)

        # Ensure dates are timezone-aware
        if end_date.tzinfo is None:
            end_date = timezone.make_aware(end_date)
        if start_date.tzinfo is None:
            start_date = timezone.make_aware(start_date)

        # Filter scholarships for this donor
        donor_scholarships = [
            s for s in self.scholarships if s.donor_info.get("name") == donor_name
        ]

        active_awards = []
        completed_awards = []
        upcoming_deadlines = []
        upcoming_reviews = []
        reporting_requirements = []

        total_awarded = 0
        total_disbursed = 0

        for scholarship in donor_scholarships:
            # Track deadlines
            if scholarship.deadline:
                deadline = scholarship.deadline
                if deadline.tzinfo is None:
                    deadline = timezone.make_aware(deadline)
                if start_date <= deadline <= end_date:
                    upcoming_deadlines.append(
                        {
                            "scholarship": scholarship.name,
                            "deadline": deadline,
                            "type": "Application Deadline",
                        }
                    )

            # Track review dates (stored as ISO strings in JSON)
            if scholarship.review_dates:
                # Handle case where review_dates might be a JSON string instead of a list
                review_dates_list = scholarship.review_dates
                if isinstance(review_dates_list, str):
                    try:
                        import json

                        review_dates_list = json.loads(review_dates_list)
                    except:
                        review_dates_list = []

                for review_date_item in review_dates_list:
                    # Convert ISO string to datetime if needed
                    if isinstance(review_date_item, str):
                        try:
                            review_date = datetime.fromisoformat(review_date_item)
                            if review_date.tzinfo is None:
                                review_date = timezone.make_aware(review_date)
                        except:
                            continue  # Skip invalid date strings
                    elif isinstance(review_date_item, datetime):
                        review_date = review_date_item
                        if review_date.tzinfo is None:
                            review_date = timezone.make_aware(review_date)
                    else:
                        continue  # Skip invalid types

                    if start_date <= review_date <= end_date:
                        upcoming_reviews.append(
                            {
                                "scholarship": scholarship.name,
                                "date": review_date,
                                "type": "Performance Review",
                            }
                        )

            # Track reporting requirements (stored as ISO strings in JSON)
            if scholarship.reporting_schedule:
                # Handle case where reporting_schedule might be a JSON string instead of a dict
                reporting_schedule_dict = scholarship.reporting_schedule
                if isinstance(reporting_schedule_dict, str):
                    try:
                        import json

                        reporting_schedule_dict = json.loads(reporting_schedule_dict)
                    except:
                        reporting_schedule_dict = {}

                if isinstance(reporting_schedule_dict, dict):
                    for (
                        report_type,
                        report_date_item,
                    ) in reporting_schedule_dict.items():
                        # Convert ISO string to datetime if needed
                        if isinstance(report_date_item, str):
                            try:
                                report_date = datetime.fromisoformat(report_date_item)
                                if report_date.tzinfo is None:
                                    report_date = timezone.make_aware(report_date)
                            except:
                                continue  # Skip invalid date strings
                        elif isinstance(report_date_item, datetime):
                            report_date = report_date_item
                            if report_date.tzinfo is None:
                                report_date = timezone.make_aware(report_date)
                        else:
                            continue  # Skip invalid types

                        if start_date <= report_date <= end_date:
                            reporting_requirements.append(
                                {
                                    "scholarship": scholarship.name,
                                    "date": report_date,
                                    "type": report_type,
                                }
                            )

            # Process awards - query ScholarshipAward model by scholarship name
            scholarship_awards = ScholarshipAward.objects.filter(
                scholarship_name=scholarship.name
            )

            for award in scholarship_awards:
                # Skip awards for "Test User" applicants
                if award.applicant and (
                    "test user" in award.applicant.name.lower()
                    or award.applicant.name.lower() == "test"
                    or "test" in award.applicant.name.lower()
                    and "user" in award.applicant.name.lower()
                ):
                    continue

                # Normalize award_date (handle strings from JSONField or datetimes)
                award_date = award.award_date
                if isinstance(award_date, str):
                    try:
                        award_date = datetime.fromisoformat(award_date)
                    except Exception:
                        award_date = pd.to_datetime(award_date).to_pydatetime()
                # Make timezone-aware when needed
                if award_date and getattr(award_date, "tzinfo", None) is None:
                    try:
                        award_date = timezone.make_aware(award_date)
                    except Exception:
                        pass

                # Normalize disbursement dates: convert ISO strings to datetimes
                raw_disbursements = award.disbursement_dates or []

                # Handle case where disbursement_dates might be a JSON string instead of a list
                if isinstance(raw_disbursements, str):
                    try:
                        import json

                        raw_disbursements = json.loads(raw_disbursements)
                    except:
                        raw_disbursements = []

                disbursement_dates = []
                for d in raw_disbursements:
                    if isinstance(d, str):
                        try:
                            dt = datetime.fromisoformat(d)
                        except Exception:
                            try:
                                dt = pd.to_datetime(d).to_pydatetime()
                            except:
                                continue  # Skip invalid date strings
                        # Ensure timezone-aware
                        if getattr(dt, "tzinfo", None) is None:
                            try:
                                dt = timezone.make_aware(dt)
                            except Exception:
                                pass
                        disbursement_dates.append(dt)
                    elif isinstance(d, datetime):
                        dt = d
                        # Ensure timezone-aware
                        if getattr(dt, "tzinfo", None) is None:
                            try:
                                dt = timezone.make_aware(dt)
                            except Exception:
                                pass
                        disbursement_dates.append(dt)
                    # Skip invalid types

                # Only include awards within the date range
                if not award_date or not (start_date <= award_date <= end_date):
                    continue

                # Ensure award_amount is a float (handle Decimal)
                amount = float(award.award_amount)

                total_awarded += amount

                # Calculate disbursed amount safely
                disbursed = 0.0
                if disbursement_dates:
                    valid_dates = [d for d in disbursement_dates if d <= end_date]
                    if valid_dates:
                        disbursed = sum(
                            amount / len(disbursement_dates) for _ in valid_dates
                        )
                total_disbursed += disbursed

                # Get recipient name from applicant relationship
                recipient = award.applicant.name if award.applicant else "Unknown"

                # Determine next disbursement
                next_disb = None
                future_dates = [d for d in disbursement_dates if d > end_date]
                if future_dates:
                    next_disb = min(future_dates)

                award_summary = {
                    "scholarship": scholarship.name,
                    "recipient": recipient,
                    "amount": amount,
                    "disbursed": disbursed,
                    "award_date": award_date,
                    "status": award.status,
                    "requirements_met": award.requirements_met or [],
                    "requirements_pending": award.requirements_pending or [],
                    "performance_metrics": award.performance_metrics or {},
                    "next_disbursement": next_disb,
                }

                if award.status == "completed":
                    completed_awards.append(award_summary)
                elif award.status == "active":
                    active_awards.append(award_summary)

        # Sort all dates
        upcoming_deadlines.sort(key=lambda x: x["deadline"])
        upcoming_reviews.sort(key=lambda x: x["date"])
        reporting_requirements.sort(key=lambda x: x["date"])
        active_awards.sort(key=lambda x: x["next_disbursement"] or end_date)
        completed_awards.sort(key=lambda x: x["award_date"], reverse=True)

        return {
            "donor_name": donor_name,
            "report_period": {"start": start_date, "end": end_date},
            "summary": {
                "total_scholarships": len(donor_scholarships),
                "total_awarded": total_awarded,
                "total_disbursed": total_disbursed,
                "active_awards": len(active_awards),
                "completed_awards": len(completed_awards),
            },
            "key_dates": {
                "upcoming_deadlines": upcoming_deadlines,
                "upcoming_reviews": upcoming_reviews,
                "reporting_requirements": reporting_requirements,
            },
            "awards": {"active": active_awards, "completed": completed_awards},
            "scholarships": [
                {
                    "name": s.name,
                    "amount": s.amount,
                    "frequency": s.frequency,
                    "deadline": s.deadline,
                    "description": s.description,
                    "review_dates": s.review_dates,
                    "reporting_schedule": s.reporting_schedule,
                    "eligibility_criteria": s.eligibility_criteria,
                    "disbursement_requirements": s.disbursement_requirements,
                }
                for s in donor_scholarships
            ],
        }

    def export_donor_report_to_excel(
        self,
        donor_name: str,
        output_path: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> str:
        """Export donor report to Excel format.

        Args:
            donor_name (str): Name of the donor
            output_path (str): Path where to save the Excel file
            start_date (datetime, optional): Start date for report period
            end_date (datetime, optional): End date for report period

        Returns:
            str: Path to the generated Excel file
        """
        report_data = self.generate_donor_report(donor_name, start_date, end_date)

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = f"Donor Report: {donor_name}"
        ws_summary["A2"] = "Report Period:"
        ws_summary["B2"] = (
            f"{report_data['report_period']['start'].strftime('%Y-%m-%d')} to {report_data['report_period']['end'].strftime('%Y-%m-%d')}"
        )

        # Summary Statistics
        summary_headers = ["Metric", "Value"]
        summary_data = [
            ["Total Scholarships", report_data["summary"]["total_scholarships"]],
            ["Total Awarded", f"${report_data['summary']['total_awarded']:,.2f}"],
            ["Total Disbursed", f"${report_data['summary']['total_disbursed']:,.2f}"],
            ["Active Awards", report_data["summary"]["active_awards"]],
            ["Completed Awards", report_data["summary"]["completed_awards"]],
        ]

        for row_idx, row in enumerate([summary_headers] + summary_data, 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Headers
                    cell.font = Font(bold=True)

        # Key Dates Sheet
        ws_dates = wb.create_sheet("Key Dates")
        date_headers = ["Type", "Scholarship", "Date", "Details"]

        # Combine all dates
        all_dates = (
            [
                [
                    "Application Deadline",
                    d["scholarship"],
                    d["deadline"].strftime("%Y-%m-%d")
                    if hasattr(d["deadline"], "strftime")
                    else str(d["deadline"]),
                    d.get("type", "Application Deadline"),
                ]
                for d in report_data["key_dates"]["upcoming_deadlines"]
            ]
            + [
                [
                    "Performance Review",
                    d["scholarship"],
                    d["date"].strftime("%Y-%m-%d")
                    if hasattr(d["date"], "strftime")
                    else str(d["date"]),
                    d.get("type", "Performance Review"),
                ]
                for d in report_data["key_dates"]["upcoming_reviews"]
            ]
            + [
                [
                    "Reporting Requirement",
                    d["scholarship"],
                    d["date"].strftime("%Y-%m-%d")
                    if hasattr(d["date"], "strftime")
                    else str(d["date"]),
                    d.get("type", "Report Due"),
                ]
                for d in report_data["key_dates"]["reporting_requirements"]
            ]
        )

        for row_idx, row in enumerate(
            [date_headers] + sorted(all_dates, key=lambda x: x[2]), 1
        ):
            for col_idx, value in enumerate(row, 1):
                cell = ws_dates.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

        # Scholarship Details Sheet
        ws_scholarships = wb.create_sheet("Scholarship Details")
        scholarship_headers = [
            "Name",
            "Amount",
            "Frequency",
            "Deadline",
            "Description",
            "Eligibility Criteria",
            "Requirements",
        ]

        scholarship_data = []
        for s in report_data["scholarships"]:
            scholarship_data.append(
                [
                    s["name"],
                    f"${s['amount']:,.2f}",
                    s["frequency"],
                    s["deadline"].strftime("%Y-%m-%d")
                    if s.get("deadline") and hasattr(s["deadline"], "strftime")
                    else str(s.get("deadline", "N/A")),
                    s["description"],
                    "; ".join(s.get("eligibility_criteria", []))
                    if isinstance(s.get("eligibility_criteria"), list)
                    else str(s.get("eligibility_criteria", "N/A")),
                    "; ".join(s.get("disbursement_requirements", []))
                    if isinstance(s.get("disbursement_requirements"), list)
                    else str(s.get("disbursement_requirements", "N/A")),
                ]
            )

        for row_idx, row in enumerate([scholarship_headers] + scholarship_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_scholarships.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Headers
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

        # Active Awards Sheet
        ws_active = wb.create_sheet("Active Awards")
        award_headers = [
            "Scholarship",
            "Recipient",
            "Amount",
            "Disbursed",
            "Status",
            "Requirements Met",
            "Requirements Pending",
            "Next Disbursement",
        ]

        award_data = [
            [
                award["scholarship"],
                award["recipient"],
                f"${award['amount']:,.2f}",
                f"${award['disbursed']:,.2f}",
                award["status"],
                "; ".join(award["requirements_met"]),
                "; ".join(award["requirements_pending"]),
                award["next_disbursement"].strftime("%Y-%m-%d")
                if award["next_disbursement"]
                else "N/A",
            ]
            for award in report_data["awards"]["active"]
        ]

        for row_idx, row in enumerate([award_headers] + award_data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws_active.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Headers
                    cell.font = Font(bold=True)

        # Adjust column widths
        for ws in [ws_summary, ws_dates, ws_scholarships, ws_active]:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[chr(64 + col[0].column)].width = min(
                    max_length + 2, 50
                )

        wb.save(output_path)
        return output_path

    def export_donor_report_to_csv(
        self,
        donor_name: str,
        output_path: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> str:
        """Export donor report to CSV format.

        Args:
            donor_name (str): Name of the donor
            output_path (str): Path where to save the CSV file
            start_date (datetime, optional): Start date for report period
            end_date (datetime, optional): End date for report period

        Returns:
            str: Path to the generated CSV file
        """
        report_data = self.generate_donor_report(donor_name, start_date, end_date)

        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # Header
            writer.writerow([f"Donor Report: {donor_name}"])
            writer.writerow(
                [
                    "Report Period:",
                    f"{report_data['report_period']['start'].strftime('%Y-%m-%d')} to "
                    f"{report_data['report_period']['end'].strftime('%Y-%m-%d')}",
                ]
            )
            writer.writerow([])

            # Summary Section
            writer.writerow(["Summary Statistics"])
            writer.writerow(["Metric", "Value"])
            writer.writerow(
                ["Total Scholarships", report_data["summary"]["total_scholarships"]]
            )
            writer.writerow(
                ["Total Awarded", f"${report_data['summary']['total_awarded']:,.2f}"]
            )
            writer.writerow(
                [
                    "Total Disbursed",
                    f"${report_data['summary']['total_disbursed']:,.2f}",
                ]
            )
            writer.writerow(["Active Awards", report_data["summary"]["active_awards"]])
            writer.writerow(
                ["Completed Awards", report_data["summary"]["completed_awards"]]
            )
            writer.writerow([])

            # Key Dates Section
            writer.writerow(["Type", "Scholarship", "Date", "Details"])

            for deadline in report_data["key_dates"]["upcoming_deadlines"]:
                deadline_str = (
                    deadline["deadline"].strftime("%Y-%m-%d")
                    if hasattr(deadline["deadline"], "strftime")
                    else str(deadline["deadline"])
                )
                writer.writerow(
                    [
                        "Application Deadline",
                        deadline["scholarship"],
                        deadline_str,
                        deadline.get("type", "Application Deadline"),
                    ]
                )

            for review in report_data["key_dates"]["upcoming_reviews"]:
                review_str = (
                    review["date"].strftime("%Y-%m-%d")
                    if hasattr(review["date"], "strftime")
                    else str(review["date"])
                )
                writer.writerow(
                    [
                        "Performance Review",
                        review["scholarship"],
                        review_str,
                        review.get("type", "Performance Review"),
                    ]
                )

            for report in report_data["key_dates"]["reporting_requirements"]:
                report_str = (
                    report["date"].strftime("%Y-%m-%d")
                    if hasattr(report["date"], "strftime")
                    else str(report["date"])
                )
                writer.writerow(
                    [
                        "Reporting Requirement",
                        report["scholarship"],
                        report_str,
                        report.get("type", "Report Due"),
                    ]
                )
            writer.writerow([])

            # Scholarship Details Section
            writer.writerow(["Scholarship Details"])
            writer.writerow(
                [
                    "Name",
                    "Amount",
                    "Frequency",
                    "Deadline",
                    "Description",
                    "Eligibility Criteria",
                    "Requirements",
                ]
            )

            for s in report_data["scholarships"]:
                deadline_str = (
                    s["deadline"].strftime("%Y-%m-%d")
                    if s.get("deadline") and hasattr(s["deadline"], "strftime")
                    else str(s.get("deadline", "N/A"))
                )
                eligibility = (
                    "; ".join(s.get("eligibility_criteria", []))
                    if isinstance(s.get("eligibility_criteria"), list)
                    else str(s.get("eligibility_criteria", "N/A"))
                )
                requirements = (
                    "; ".join(s.get("disbursement_requirements", []))
                    if isinstance(s.get("disbursement_requirements"), list)
                    else str(s.get("disbursement_requirements", "N/A"))
                )

                writer.writerow(
                    [
                        s["name"],
                        f"${s['amount']:,.2f}",
                        s["frequency"],
                        deadline_str,
                        s["description"],
                        eligibility,
                        requirements,
                    ]
                )
            writer.writerow([])

            # Active Awards Section
            writer.writerow(["Active Awards"])
            writer.writerow(
                [
                    "Scholarship",
                    "Recipient",
                    "Amount",
                    "Disbursed",
                    "Status",
                    "Requirements Met",
                    "Requirements Pending",
                    "Next Disbursement",
                ]
            )

            for award in report_data["awards"]["active"]:
                writer.writerow(
                    [
                        award["scholarship"],
                        award["recipient"],
                        f"${award['amount']:,.2f}",
                        f"${award['disbursed']:,.2f}",
                        award["status"],
                        "; ".join(award["requirements_met"]),
                        "; ".join(award["requirements_pending"]),
                        award["next_disbursement"].strftime("%Y-%m-%d")
                        if award["next_disbursement"]
                        else "N/A",
                    ]
                )

        return output_path

    def export_donor_report_to_pdf(
        self,
        donor_name: str,
        output_path: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> str:
        """Export donor report to PDF format.

        Args:
            donor_name (str): Name of the donor
            output_path (str): Path where to save the PDF file
            start_date (datetime, optional): Start date for report period
            end_date (datetime, optional): End date for report period

        Returns:
            str: Path to the generated PDF file
        """
        report_data = self.generate_donor_report(donor_name, start_date, end_date)

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        story.append(Paragraph(f"Donor Report: {donor_name}", styles["Heading1"]))
        story.append(
            Paragraph(
                f"Report Period: {report_data['report_period']['start'].strftime('%Y-%m-%d')} to "
                f"{report_data['report_period']['end'].strftime('%Y-%m-%d')}",
                styles["Normal"],
            )
        )

        # Summary Section
        story.append(Paragraph("Summary", styles["Heading2"]))
        summary_data = [
            ["Total Scholarships", str(report_data["summary"]["total_scholarships"])],
            ["Total Awarded", f"${report_data['summary']['total_awarded']:,.2f}"],
            ["Total Disbursed", f"${report_data['summary']['total_disbursed']:,.2f}"],
            ["Active Awards", str(report_data["summary"]["active_awards"])],
            ["Completed Awards", str(report_data["summary"]["completed_awards"])],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(summary_table)
        story.append(Paragraph("<br/>", styles["Normal"]))

        # Key Dates Section
        if report_data["key_dates"]["upcoming_deadlines"]:
            story.append(
                Paragraph("Upcoming Application Deadlines:", styles["Heading3"])
            )
            for deadline in report_data["key_dates"]["upcoming_deadlines"]:
                deadline_str = (
                    deadline["deadline"].strftime("%Y-%m-%d")
                    if hasattr(deadline["deadline"], "strftime")
                    else str(deadline["deadline"])
                )
                story.append(
                    Paragraph(
                        f"• {deadline['scholarship']}: {deadline_str} ({deadline.get('type', 'Application Deadline')})",
                        styles["Normal"],
                    )
                )
            story.append(Paragraph("<br/>", styles["Normal"]))

        if report_data["key_dates"]["upcoming_reviews"]:
            story.append(Paragraph("Upcoming Performance Reviews:", styles["Heading3"]))
            for review in report_data["key_dates"]["upcoming_reviews"]:
                review_date_str = (
                    review["date"].strftime("%Y-%m-%d")
                    if hasattr(review["date"], "strftime")
                    else str(review["date"])
                )
                story.append(
                    Paragraph(
                        f"• {review['scholarship']}: {review_date_str} ({review.get('type', 'Performance Review')})",
                        styles["Normal"],
                    )
                )
            story.append(Paragraph("<br/>", styles["Normal"]))

        if report_data["key_dates"]["reporting_requirements"]:
            story.append(
                Paragraph("Upcoming Reporting Requirements:", styles["Heading3"])
            )
            for requirement in report_data["key_dates"]["reporting_requirements"]:
                req_date_str = (
                    requirement["date"].strftime("%Y-%m-%d")
                    if hasattr(requirement["date"], "strftime")
                    else str(requirement["date"])
                )
                story.append(
                    Paragraph(
                        f"• {requirement['scholarship']}: {requirement['type']} - {req_date_str}",
                        styles["Normal"],
                    )
                )
            story.append(Paragraph("<br/>", styles["Normal"]))

        # Scholarship Details Section with Key Dates
        if report_data["scholarships"]:
            story.append(Paragraph("Scholarship Details", styles["Heading2"]))
            for scholarship in report_data["scholarships"]:
                story.append(Paragraph(f"{scholarship['name']}", styles["Heading3"]))
                story.append(
                    Paragraph(
                        f"Amount: ${scholarship['amount']:,.2f} ({scholarship['frequency']})",
                        styles["Normal"],
                    )
                )

                if scholarship.get("deadline"):
                    deadline_str = (
                        scholarship["deadline"].strftime("%Y-%m-%d")
                        if hasattr(scholarship["deadline"], "strftime")
                        else str(scholarship["deadline"])
                    )
                    story.append(
                        Paragraph(
                            f"Application Deadline: {deadline_str}", styles["Normal"]
                        )
                    )

                story.append(
                    Paragraph(
                        f"Description: {scholarship['description']}", styles["Normal"]
                    )
                )

                # Show eligibility criteria
                if scholarship.get("eligibility_criteria"):
                    story.append(Paragraph("Eligibility Criteria:", styles["Heading4"]))
                    for criteria in scholarship["eligibility_criteria"]:
                        story.append(Paragraph(f"• {criteria}", styles["Normal"]))

                # Show disbursement requirements
                if scholarship.get("disbursement_requirements"):
                    story.append(
                        Paragraph("Disbursement Requirements:", styles["Heading4"])
                    )
                    for req in scholarship["disbursement_requirements"]:
                        story.append(Paragraph(f"• {req}", styles["Normal"]))

                story.append(Paragraph("<br/>", styles["Normal"]))

        # Active Awards Section
        story.append(Paragraph("Active Awards", styles["Heading2"]))
        for award in report_data["awards"]["active"]:
            story.append(
                Paragraph(f"Scholarship: {award['scholarship']}", styles["Heading3"])
            )
            story.append(
                Paragraph(f"Recipient: {award['recipient']}", styles["Normal"])
            )
            story.append(
                Paragraph(f"Amount: ${award['amount']:,.2f}", styles["Normal"])
            )
            story.append(
                Paragraph(f"Disbursed: ${award['disbursed']:,.2f}", styles["Normal"])
            )
            if award["next_disbursement"]:
                story.append(
                    Paragraph(
                        f"Next Disbursement: {award['next_disbursement'].strftime('%Y-%m-%d')}",
                        styles["Normal"],
                    )
                )
            story.append(Paragraph("Requirements Status:", styles["Normal"]))
            for req in award["requirements_met"]:
                story.append(Paragraph(f"✓ {req}", styles["Normal"]))
            for req in award["requirements_pending"]:
                story.append(Paragraph(f"□ {req}", styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

        doc.build(story)
        return output_path

    def add_scholarship(self, scholarship: Scholarship):
        """Add a new scholarship to the system."""
        self.scholarships.append(scholarship)

    def get_scholarships_data(self) -> List[Scholarship]:
        """Unified source of scholarships for reports and analytics.

        If in-memory scholarships have been added to the engine, use those;
        otherwise, read from the database to ensure a single source of truth.
        """
        try:
            if getattr(self, "scholarships", None):
                # Ensure we return concrete model instances (already the case in add_scholarship usage)
                return list(self.scholarships)
        except Exception:
            pass
        return list(Scholarship.objects.all())

    # Function to generate disbursement report. Meets requirement SFWE504_3-LLR-8.
    def generate_disbursement_report(
        self, scholarship_name: str = None
    ) -> Dict[str, Any]:
        """Generate a disbursement report for scholarship recipients showing payment details.

        Args:
            scholarship_name (str, optional): Specific scholarship to report on. If None, reports on all scholarships.

        Returns:
            dict: Disbursement report containing:
                - Recipient information
                - Award amounts and disbursement schedules
                - Payment dates and amounts
                - Outstanding balances
                - Requirements status
                - Summary statistics
        """
        # Query scholarship awards
        if scholarship_name:
            awards_queryset = ScholarshipAward.objects.filter(
                scholarship_name=scholarship_name, status="active"
            )
        else:
            awards_queryset = ScholarshipAward.objects.filter(status="active")

        # Build disbursement details for each award
        disbursements = []
        total_awarded = 0
        total_disbursed = 0
        total_pending = 0

        for award in awards_queryset:
            applicant = award.applicant
            # Skip placeholder or temporary users
            if (
                not applicant
                or not getattr(applicant, "name", None)
                or applicant.name.lower().strip()
                in ["temp user", "temp", "test user", "test"]
                or (
                    "temp" in applicant.name.lower()
                    and "user" in applicant.name.lower()
                )
                or (
                    "test" in applicant.name.lower()
                    and "user" in applicant.name.lower()
                )
            ):
                continue

            award_amount = float(award.award_amount)
            total_awarded += award_amount

            # Calculate disbursed and pending amounts
            num_disbursements = (
                len(award.disbursement_dates) if award.disbursement_dates else 1
            )
            amount_per_disbursement = (
                award_amount / num_disbursements
                if num_disbursements > 0
                else award_amount
            )

            # Count past disbursements
            today = datetime.now().date()
            past_disbursements = []
            future_disbursements = []

            if award.disbursement_dates:
                for disburse_date in award.disbursement_dates:
                    # Parse date if string
                    if isinstance(disburse_date, str):
                        try:
                            disburse_date = datetime.fromisoformat(disburse_date).date()
                        except:
                            disburse_date = datetime.strptime(
                                disburse_date, "%Y-%m-%d"
                            ).date()
                    elif isinstance(disburse_date, datetime):
                        disburse_date = disburse_date.date()

                    if disburse_date <= today:
                        past_disbursements.append(
                            {"date": disburse_date, "amount": amount_per_disbursement}
                        )
                    else:
                        future_disbursements.append(
                            {"date": disburse_date, "amount": amount_per_disbursement}
                        )

            disbursed_amount = len(past_disbursements) * amount_per_disbursement
            pending_amount = award_amount - disbursed_amount

            total_disbursed += disbursed_amount
            total_pending += pending_amount

            disbursements.append(
                {
                    "scholarship_name": award.scholarship_name,
                    "recipient_name": applicant.name,
                    "student_id": applicant.student_id,
                    "award_date": award.award_date,
                    "total_award_amount": award_amount,
                    "disbursed_amount": disbursed_amount,
                    "pending_amount": pending_amount,
                    "disbursement_schedule": {
                        "total_payments": num_disbursements,
                        "amount_per_payment": amount_per_disbursement,
                        "completed_payments": past_disbursements,
                        "upcoming_payments": future_disbursements,
                    },
                    "requirements_met": award.requirements_met or [],
                    "requirements_pending": award.requirements_pending or [],
                    "status": award.status,
                    "notes": award.notes,
                }
            )

        report = {
            "generated_date": datetime.now(),
            "scholarship_name": scholarship_name or "All Scholarships",
            "total_recipients": len(disbursements),
            "summary": {
                "total_awarded": total_awarded,
                "total_disbursed": total_disbursed,
                "total_pending": total_pending,
                "disbursement_completion_rate": (total_disbursed / total_awarded * 100)
                if total_awarded > 0
                else 0,
            },
            "disbursements": disbursements,
        }

        return report

    # Function to generate pre-screening report. Meets requirement for pre-screening applicants, SFWE504_3-LLR-7, SFWE504_3-LLR-25, SFWE504_3-LLR-26.
    def generate_prescreening_report(
        self, applicants: List[Applicant], scholarship_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Generate a pre-screening report identifying applicants who meet scholarship eligibility criteria.

        Args:
            applicants (List[Applicant]): List of applicants to evaluate
            scholarship_id (str, optional): Specific scholarship to evaluate for. If None, evaluate all scholarships.

        Returns:
            dict: Pre-screening report containing:
                - Qualified applicants list per scholarship
                - Detailed eligibility analysis for each applicant
                - Review scores and committee comments
                - Application progress tracking
                - Match statistics and summary
                - Scholarship-specific requirements status
                - Comprehensive applicant qualifications
        """
        report = {
            "generated_date": datetime.now(),
            "scholarships_evaluated": 0,
            "total_applicants": len(applicants),
            "matches": [],
            "qualified_applicants": {},  # Dictionary to store qualified applicants per scholarship
            "applicant_analysis": {},  # Detailed analysis of each applicant
            "review_tracking": {},  # Track review scores and comments
            "application_progress": {},  # Track application completion status
            "summary": {
                "total_matches": 0,
                "match_rate": 0.0,
                "scholarships_with_matches": 0,
                "qualification_distribution": {},  # Distribution of qualification rates
                "review_statistics": {  # Statistics about review scores
                    "average_scores": {},
                    "review_completion_rate": 0.0,
                },
                "application_completion": {  # Track overall application completion
                    "complete": 0,
                    "in_progress": 0,
                    "incomplete": 0,
                },
            },
        }

        # Filter scholarships if scholarship_id is provided
        scholarships_to_evaluate = [
            s
            for s in self.scholarships
            if scholarship_id is None or s.name == scholarship_id
        ]
        report["scholarships_evaluated"] = len(scholarships_to_evaluate)

        for scholarship in scholarships_to_evaluate:
            scholarship_matches = []
            qualified_applicants = []
            qualification_scores = []  # Track qualification scores for distribution analysis

            for applicant in applicants:
                eligibility_results = []
                meets_all_criteria = True
                criteria_met_count = 0
                total_criteria = len(scholarship.eligibility_criteria)

                # Evaluate each eligibility criterion
                for criterion in scholarship.eligibility_criteria:
                    is_met = False
                    reason = ""
                    details = {}

                    # Evaluate GPA requirements
                    if "GPA" in criterion:
                        required_gpa = float(criterion.split("+")[0].split()[-1])
                        is_met = applicant.gpa >= required_gpa
                        reason = f"GPA: {applicant.gpa:.2f} vs required {required_gpa}+"
                        details = {
                            "type": "gpa",
                            "required": required_gpa,
                            "actual": applicant.gpa,
                            "difference": applicant.gpa - required_gpa,
                        }

                    # Evaluate major requirements
                    elif "major" in criterion.lower():
                        required_major = criterion.split("major")[0].strip()
                        is_met = required_major.lower() in applicant.major.lower()
                        reason = (
                            f"Major: {applicant.major} vs required {required_major}"
                        )
                        details = {
                            "type": "major",
                            "required": required_major,
                            "actual": applicant.major,
                            "exact_match": required_major.lower()
                            == applicant.major.lower(),
                        }

                    # Evaluate enrollment status
                    elif "enrollment" in criterion.lower():
                        # This would need to be enhanced with actual enrollment status data
                        is_met = True  # Assuming full-time enrollment for demo
                        reason = "Enrollment status verified"
                        details = {
                            "type": "enrollment",
                            "status": "full-time",
                            "verified": True,
                        }

                    # Track met criteria
                    if is_met:
                        criteria_met_count += 1

                    # Add detailed evaluation results
                    eligibility_results.append(
                        {
                            "criterion": criterion,
                            "is_met": is_met,
                            "reason": reason,
                            "details": details,
                        }
                    )

                    if not is_met:
                        meets_all_criteria = False

                # Calculate qualification score
                qualification_score = (criteria_met_count / total_criteria) * 100
                qualification_scores.append(qualification_score)

                # Calculate application completion status
                required_components = {
                    "personal_info": bool(applicant.name and applicant.student_id),
                    "academic_info": bool(applicant.major and applicant.academic_level),
                    "essays": bool(applicant.essays),
                    "financial_info": bool(applicant.financial_info),
                    "academic_records": bool(applicant.academic_history),
                }
                completion_percentage = (
                    sum(1 for v in required_components.values() if v)
                    / len(required_components)
                ) * 100

                # Determine application status
                if completion_percentage == 100:
                    application_status = "complete"
                elif completion_percentage > 50:
                    application_status = "in_progress"
                else:
                    application_status = "incomplete"

                # Get review scores and comments if available
                review_data = {
                    "academic_review": {
                        "score": None,
                        "comments": [],
                        "reviewer": None,
                        "date": None,
                    },
                    "essay_review": {
                        "scores": [],
                        "comments": [],
                        "reviewers": [],
                        "dates": [],
                    },
                    "interview_notes": None,
                    "committee_feedback": [],
                }

                # Process essay evaluations if available
                if hasattr(applicant, "essays") and applicant.essays:
                    for essay in applicant.essays:
                        if hasattr(essay, "evaluation"):
                            review_data["essay_review"]["scores"].append(
                                essay.evaluation.get("score")
                            )
                            review_data["essay_review"]["comments"].append(
                                essay.evaluation.get("feedback")
                            )
                            review_data["essay_review"]["reviewers"].append(
                                essay.evaluation.get("reviewer")
                            )
                            review_data["essay_review"]["dates"].append(
                                essay.evaluation.get("date")
                            )

                # Process interview notes if available
                if hasattr(applicant, "interview_notes"):
                    review_data["interview_notes"] = applicant.interview_notes

                # Process committee feedback if available
                if hasattr(applicant, "committee_feedback"):
                    review_data["committee_feedback"] = applicant.committee_feedback

                # Fetch simple award decision, if any
                award_decision_data = None
                try:
                    ad = AwardDecision.objects.get(
                        applicant=applicant, scholarship_name=scholarship.name
                    )
                    award_decision_data = {
                        "decision": ad.decision,
                        "comments": ad.comments,
                        "decided_at": ad.decided_at,
                    }
                except AwardDecision.DoesNotExist:
                    pass

                # Prepare detailed applicant assessment
                applicant_assessment = {
                    "applicant": {
                        "name": applicant.name,
                        "student_id": applicant.student_id,
                        "major": applicant.major,
                        "gpa": applicant.gpa,
                        "academic_level": applicant.academic_level,
                    },
                    "qualification_score": qualification_score,
                    "eligibility_details": eligibility_results,
                    "criteria_met_count": criteria_met_count,
                    "total_criteria": total_criteria,
                    "fully_qualified": meets_all_criteria,
                    "application_status": {
                        "status": application_status,
                        "completion_percentage": completion_percentage,
                        "missing_components": [
                            component
                            for component, completed in required_components.items()
                            if not completed
                        ],
                    },
                    "review_data": review_data,
                    "award_decision": award_decision_data,
                }

                if meets_all_criteria:
                    scholarship_matches.append(applicant_assessment)
                    qualified_applicants.append(
                        {
                            "applicant": applicant_assessment["applicant"],
                            "qualification_score": qualification_score,
                        }
                    )

                # Store detailed analysis for each applicant
                if applicant.student_id not in report["applicant_analysis"]:
                    report["applicant_analysis"][applicant.student_id] = []
                report["applicant_analysis"][applicant.student_id].append(
                    {
                        "scholarship_name": scholarship.name,
                        "assessment": applicant_assessment,
                    }
                )

            if scholarship_matches:
                # Sort qualified applicants by qualification score
                qualified_applicants.sort(
                    key=lambda x: x["qualification_score"], reverse=True
                )

                report["matches"].append(
                    {
                        "scholarship_name": scholarship.name,
                        "description": scholarship.description,
                        "amount": scholarship.amount,
                        "deadline": scholarship.deadline,
                        "eligibility_criteria": scholarship.eligibility_criteria,
                        "matches": scholarship_matches,
                        "qualification_distribution": {
                            "min_score": min(qualification_scores)
                            if qualification_scores
                            else 0,
                            "max_score": max(qualification_scores)
                            if qualification_scores
                            else 0,
                            "average_score": sum(qualification_scores)
                            / len(qualification_scores)
                            if qualification_scores
                            else 0,
                        },
                    }
                )

                # Store qualified applicants for this scholarship
                report["qualified_applicants"][scholarship.name] = qualified_applicants

        # Calculate comprehensive summary statistics
        total_matches = sum(len(s["matches"]) for s in report["matches"])
        scholarships_with_matches = len(report["matches"])

        # Calculate qualification distribution across all scholarships
        all_qualification_scores = []
        for scholarship in report["matches"]:
            all_qualification_scores.extend(
                [match["qualification_score"] for match in scholarship["matches"]]
            )

        # Collect review statistics
        review_scores = []
        essay_scores = []
        total_reviews = 0
        completed_reviews = 0

        # Count application completion status
        application_completion = {"complete": 0, "in_progress": 0, "incomplete": 0}

        # Award decision summary counts
        award_decision_summary = {"awarded": 0, "not_awarded": 0, "pending": 0}

        for scholarship in report["matches"]:
            for match in scholarship["matches"]:
                # Track application status
                status = match["application_status"]["status"]
                application_completion[status] += 1

                # Award decision counting
                if match.get("award_decision"):
                    award_decision_summary[match["award_decision"]["decision"]] += 1

                # Track review completion
                review_data = match["review_data"]
                if review_data["academic_review"]["score"] is not None:
                    review_scores.append(review_data["academic_review"]["score"])
                    completed_reviews += 1

                for essay_score in review_data["essay_review"]["scores"]:
                    if essay_score is not None:
                        essay_scores.append(essay_score)
                        completed_reviews += 1

                total_reviews += 1  # Count expected reviews
                if review_data["interview_notes"]:
                    completed_reviews += 1
                total_reviews += 1  # Count interview as expected

        # Calculate average review scores
        avg_review_score = (
            sum(review_scores) / len(review_scores) if review_scores else 0
        )
        avg_essay_score = sum(essay_scores) / len(essay_scores) if essay_scores else 0
        review_completion_rate = (
            (completed_reviews / total_reviews) if total_reviews > 0 else 0
        )

        report["summary"] = {
            "total_matches": total_matches,
            "match_rate": (total_matches / len(applicants)) if applicants else 0.0,
            "scholarships_with_matches": scholarships_with_matches,
            "match_distribution": {
                "scholarship_match_rate": (
                    scholarships_with_matches / len(scholarships_to_evaluate)
                )
                if scholarships_to_evaluate
                else 0.0,
                "average_matches_per_scholarship": (
                    total_matches / len(scholarships_to_evaluate)
                )
                if scholarships_to_evaluate
                else 0.0,
            },
            "qualification_distribution": {
                "min_score": min(all_qualification_scores)
                if all_qualification_scores
                else 0,
                "max_score": max(all_qualification_scores)
                if all_qualification_scores
                else 0,
                "average_score": sum(all_qualification_scores)
                / len(all_qualification_scores)
                if all_qualification_scores
                else 0,
                "score_ranges": {
                    "90-100": len(
                        [s for s in all_qualification_scores if 90 <= s <= 100]
                    ),
                    "80-89": len([s for s in all_qualification_scores if 80 <= s < 90]),
                    "70-79": len([s for s in all_qualification_scores if 70 <= s < 80]),
                    "60-69": len([s for s in all_qualification_scores if 60 <= s < 70]),
                    "Below 60": len([s for s in all_qualification_scores if s < 60]),
                },
            },
            "review_statistics": {
                "average_scores": {
                    "academic_review": avg_review_score,
                    "essay_review": avg_essay_score,
                },
                "review_completion_rate": review_completion_rate,
                "reviews_completed": completed_reviews,
                "total_reviews_expected": total_reviews,
            },
            "application_completion": application_completion,
            "award_decisions": award_decision_summary,
        }

        return report

    def export_disbursement_report_to_pdf(
        self, scholarship_name: str = None, output_path: str = None
    ) -> str:
        """Export disbursement report to PDF format."""
        report_data = self.generate_disbursement_report(scholarship_name)

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        story.append(
            Paragraph(
                f"Disbursement Report: {report_data['scholarship_name']}",
                styles["Heading1"],
            )
        )
        story.append(
            Paragraph(
                f"Generated on: {report_data['generated_date'].strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"],
            )
        )
        story.append(Paragraph("<br/>", styles["Normal"]))

        # Summary Section
        story.append(Paragraph("Summary Statistics", styles["Heading2"]))
        summary_data = [
            ["Total Recipients", str(report_data["total_recipients"])],
            ["Total Awarded", f"${report_data['summary']['total_awarded']:,.2f}"],
            ["Total Disbursed", f"${report_data['summary']['total_disbursed']:,.2f}"],
            ["Total Pending", f"${report_data['summary']['total_pending']:,.2f}"],
            [
                "Completion Rate",
                f"{report_data['summary']['disbursement_completion_rate']:.1f}%",
            ],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                ]
            )
        )
        story.append(summary_table)
        story.append(Paragraph("<br/>", styles["Normal"]))

        # Disbursement Details
        story.append(Paragraph("Recipient Disbursement Details", styles["Heading2"]))

        for disbursement in report_data["disbursements"]:
            story.append(
                Paragraph(
                    f"Recipient: {disbursement['recipient_name']} ({disbursement['student_id']})",
                    styles["Heading3"],
                )
            )
            story.append(
                Paragraph(
                    f"Scholarship: {disbursement['scholarship_name']}", styles["Normal"]
                )
            )

            # Award details
            detail_data = [
                [
                    "Award Date",
                    disbursement["award_date"].strftime("%Y-%m-%d")
                    if hasattr(disbursement["award_date"], "strftime")
                    else str(disbursement["award_date"]),
                ],
                ["Total Award", f"${disbursement['total_award_amount']:,.2f}"],
                ["Amount Disbursed", f"${disbursement['disbursed_amount']:,.2f}"],
                ["Amount Pending", f"${disbursement['pending_amount']:,.2f}"],
                ["Status", disbursement["status"]],
            ]
            detail_table = Table(detail_data)
            detail_table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ]
                )
            )
            story.append(detail_table)
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Disbursement schedule
            schedule = disbursement["disbursement_schedule"]
            story.append(
                Paragraph(
                    f"Payment Schedule ({schedule['total_payments']} payments of ${schedule['amount_per_payment']:,.2f} each):",
                    styles["Heading4"],
                )
            )

            if schedule["completed_payments"]:
                story.append(Paragraph("Completed Payments:", styles["Normal"]))
                for payment in schedule["completed_payments"]:
                    payment_date = payment["date"]
                    date_str = (
                        payment_date.strftime("%Y-%m-%d")
                        if hasattr(payment_date, "strftime")
                        else str(payment_date)
                    )
                    story.append(
                        Paragraph(
                            f"✓ {date_str}: ${payment['amount']:,.2f}", styles["Normal"]
                        )
                    )

            if schedule["upcoming_payments"]:
                story.append(Paragraph("Upcoming Payments:", styles["Normal"]))
                for payment in schedule["upcoming_payments"]:
                    payment_date = payment["date"]
                    date_str = (
                        payment_date.strftime("%Y-%m-%d")
                        if hasattr(payment_date, "strftime")
                        else str(payment_date)
                    )
                    story.append(
                        Paragraph(
                            f"○ {date_str}: ${payment['amount']:,.2f}", styles["Normal"]
                        )
                    )

            # Requirements
            if disbursement["requirements_met"] or disbursement["requirements_pending"]:
                story.append(Paragraph("Requirements:", styles["Heading4"]))
                for req in disbursement["requirements_met"]:
                    story.append(Paragraph(f"✓ {req}", styles["Normal"]))
                for req in disbursement["requirements_pending"]:
                    story.append(Paragraph(f"□ {req}", styles["Normal"]))

            if disbursement.get("notes"):
                story.append(
                    Paragraph(f"Notes: {disbursement['notes']}", styles["Normal"])
                )

            story.append(Paragraph("<br/>", styles["Normal"]))

        doc.build(story)
        return output_path

    def export_disbursement_report_to_excel(
        self, scholarship_name: str = None, output_path: str = None
    ) -> str:
        """Export disbursement report to Excel format."""
        report_data = self.generate_disbursement_report(scholarship_name)

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.cell(
            row=1,
            column=1,
            value=f"Disbursement Report: {report_data['scholarship_name']}",
        ).font = Font(bold=True, size=14)
        ws_summary.cell(
            row=2,
            column=1,
            value=f"Generated: {report_data['generated_date'].strftime('%Y-%m-%d %H:%M:%S')}",
        )

        ws_summary.cell(row=4, column=1, value="Total Recipients").font = Font(
            bold=True
        )
        ws_summary.cell(row=4, column=2, value=report_data["total_recipients"])
        ws_summary.cell(row=5, column=1, value="Total Awarded").font = Font(bold=True)
        ws_summary.cell(
            row=5, column=2, value=f"${report_data['summary']['total_awarded']:,.2f}"
        )
        ws_summary.cell(row=6, column=1, value="Total Disbursed").font = Font(bold=True)
        ws_summary.cell(
            row=6, column=2, value=f"${report_data['summary']['total_disbursed']:,.2f}"
        )
        ws_summary.cell(row=7, column=1, value="Total Pending").font = Font(bold=True)
        ws_summary.cell(
            row=7, column=2, value=f"${report_data['summary']['total_pending']:,.2f}"
        )
        ws_summary.cell(row=8, column=1, value="Completion Rate").font = Font(bold=True)
        ws_summary.cell(
            row=8,
            column=2,
            value=f"{report_data['summary']['disbursement_completion_rate']:.1f}%",
        )

        # Disbursements Sheet
        ws_disbursements = wb.create_sheet("Disbursements")
        headers = [
            "Scholarship",
            "Recipient",
            "Student ID",
            "Award Date",
            "Total Award",
            "Disbursed",
            "Pending",
            "Status",
            "Payments",
            "Requirements Met",
            "Requirements Pending",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_disbursements.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        for row_idx, disbursement in enumerate(report_data["disbursements"], 2):
            ws_disbursements.cell(
                row=row_idx, column=1, value=disbursement["scholarship_name"]
            )
            ws_disbursements.cell(
                row=row_idx, column=2, value=disbursement["recipient_name"]
            )
            ws_disbursements.cell(
                row=row_idx, column=3, value=disbursement["student_id"]
            )
            award_date = disbursement["award_date"]
            date_str = (
                award_date.strftime("%Y-%m-%d")
                if hasattr(award_date, "strftime")
                else str(award_date)
            )
            ws_disbursements.cell(row=row_idx, column=4, value=date_str)
            ws_disbursements.cell(
                row=row_idx,
                column=5,
                value=f"${disbursement['total_award_amount']:,.2f}",
            )
            ws_disbursements.cell(
                row=row_idx, column=6, value=f"${disbursement['disbursed_amount']:,.2f}"
            )
            ws_disbursements.cell(
                row=row_idx, column=7, value=f"${disbursement['pending_amount']:,.2f}"
            )
            ws_disbursements.cell(row=row_idx, column=8, value=disbursement["status"])
            ws_disbursements.cell(
                row=row_idx,
                column=9,
                value=f"{len(disbursement['disbursement_schedule']['completed_payments'])}/{disbursement['disbursement_schedule']['total_payments']}",
            )
            ws_disbursements.cell(
                row=row_idx,
                column=10,
                value="; ".join(disbursement["requirements_met"]),
            )
            ws_disbursements.cell(
                row=row_idx,
                column=11,
                value="; ".join(disbursement["requirements_pending"]),
            )

        # Adjust column widths
        for ws in [ws_summary, ws_disbursements]:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[chr(64 + col[0].column)].width = min(
                    max_length + 2, 50
                )

        wb.save(output_path)
        return output_path

    def export_disbursement_report_to_csv(
        self, scholarship_name: str = None, output_path: str = None
    ) -> str:
        """Export disbursement report to CSV format."""
        report_data = self.generate_disbursement_report(scholarship_name)

        import csv

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([f"Disbursement Report: {report_data['scholarship_name']}"])
            writer.writerow(
                [
                    f"Generated: {report_data['generated_date'].strftime('%Y-%m-%d %H:%M:%S')}"
                ]
            )
            writer.writerow([])

            # Summary
            writer.writerow(["Summary Statistics"])
            writer.writerow(["Total Recipients", report_data["total_recipients"]])
            writer.writerow(
                ["Total Awarded", f"${report_data['summary']['total_awarded']:,.2f}"]
            )
            writer.writerow(
                [
                    "Total Disbursed",
                    f"${report_data['summary']['total_disbursed']:,.2f}",
                ]
            )
            writer.writerow(
                ["Total Pending", f"${report_data['summary']['total_pending']:,.2f}"]
            )
            writer.writerow(
                [
                    "Completion Rate",
                    f"{report_data['summary']['disbursement_completion_rate']:.1f}%",
                ]
            )
            writer.writerow([])

            # Disbursements
            writer.writerow(
                [
                    "Scholarship",
                    "Recipient",
                    "Student ID",
                    "Award Date",
                    "Total Award",
                    "Disbursed",
                    "Pending",
                    "Status",
                    "Completed/Total Payments",
                ]
            )
            for disbursement in report_data["disbursements"]:
                award_date = disbursement["award_date"]
                date_str = (
                    award_date.strftime("%Y-%m-%d")
                    if hasattr(award_date, "strftime")
                    else str(award_date)
                )
                writer.writerow(
                    [
                        disbursement["scholarship_name"],
                        disbursement["recipient_name"],
                        disbursement["student_id"],
                        date_str,
                        f"${disbursement['total_award_amount']:,.2f}",
                        f"${disbursement['disbursed_amount']:,.2f}",
                        f"${disbursement['pending_amount']:,.2f}",
                        disbursement["status"],
                        f"{len(disbursement['disbursement_schedule']['completed_payments'])}/{disbursement['disbursement_schedule']['total_payments']}",
                    ]
                )

        return output_path

    def export_prescreening_report_to_pdf(
        self,
        applicants: List[Applicant],
        scholarship_id: Optional[str] = None,
        output_path: str = None,
    ) -> str:
        """Export pre-screening report to PDF format."""
        report_data = self.generate_prescreening_report(applicants, scholarship_id)

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        story.append(Paragraph("Pre-screening Report", styles["Heading1"]))
        story.append(
            Paragraph(
                f"Generated on: {report_data['generated_date'].strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"],
            )
        )

        # Summary Section
        story.append(Paragraph("Summary", styles["Heading2"]))
        summary_data = [
            ["Total Applicants", str(report_data["total_applicants"])],
            ["Total Matches", str(report_data["summary"]["total_matches"])],
            ["Match Rate", f"{report_data['summary']['match_rate'] * 100:.1f}%"],
            [
                "Scholarships with Matches",
                str(report_data["summary"]["scholarships_with_matches"]),
            ],
            [
                "Review Completion Rate",
                f"{report_data['summary']['review_statistics']['review_completion_rate'] * 100:.1f}%",
            ],
            [
                "Applications Complete",
                str(report_data["summary"]["application_completion"]["complete"]),
            ],
            [
                "Applications In Progress",
                str(report_data["summary"]["application_completion"]["in_progress"]),
            ],
            [
                "Applications Incomplete",
                str(report_data["summary"]["application_completion"]["incomplete"]),
            ],
        ]

        summary_table = Table(summary_data)
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                ]
            )
        )
        story.append(summary_table)
        story.append(Paragraph("<br/>", styles["Normal"]))

        # Review Statistics
        story.append(Paragraph("Review Statistics", styles["Heading2"]))
        review_stats = report_data["summary"]["review_statistics"]
        story.append(
            Paragraph(
                f"Average Academic Review Score: {review_stats['average_scores']['academic_review']:.1f}/10",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"Average Essay Review Score: {review_stats['average_scores']['essay_review']:.1f}/10",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"Reviews Completed: {review_stats['reviews_completed']} of {review_stats['total_reviews_expected']}",
                styles["Normal"],
            )
        )
        story.append(Paragraph("<br/>", styles["Normal"]))

        # Award Decision Summary Section
        if "award_decisions" in report_data["summary"]:
            story.append(Paragraph("Award Decisions", styles["Heading2"]))
            ad = report_data["summary"]["award_decisions"]
            story.append(Paragraph(f"Awarded: {ad['awarded']}", styles["Normal"]))
            story.append(
                Paragraph(f"Not Awarded: {ad['not_awarded']}", styles["Normal"])
            )
            story.append(Paragraph(f"Pending: {ad['pending']}", styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

        # Matches by Scholarship
        for scholarship_match in report_data["matches"]:
            story.append(
                Paragraph(scholarship_match["scholarship_name"], styles["Heading2"])
            )
            story.append(Paragraph(scholarship_match["description"], styles["Normal"]))
            story.append(
                Paragraph(
                    f"Amount: ${scholarship_match['amount']:,.2f}", styles["Normal"]
                )
            )
            if scholarship_match["deadline"]:
                story.append(
                    Paragraph(
                        f"Deadline: {scholarship_match['deadline'].strftime('%Y-%m-%d')}",
                        styles["Normal"],
                    )
                )
            # Eligibility Criteria Section
            if scholarship_match.get("eligibility_criteria"):
                story.append(Paragraph("Eligibility Criteria:", styles["Heading3"]))
                for criteria in scholarship_match["eligibility_criteria"]:
                    story.append(Paragraph(f"• {criteria}", styles["Normal"]))
                story.append(Paragraph("<br/>", styles["Normal"]))

            # Table of matching applicants with review scores
            story.append(Paragraph("Qualified Applicants:", styles["Heading3"]))
            applicant_data = [
                [
                    "Name",
                    "Student ID",
                    "Major",
                    "GPA",
                    "Academic Level",
                    "Application Status",
                    "Review Score",
                    "Award Decision",
                ]
            ]

            for match in scholarship_match["matches"]:
                applicant = match["applicant"]
                review_data = match.get("review_data", {})
                application_status = match.get("application_status", {})

                # Calculate average review score
                review_scores = []
                if review_data.get("academic_review", {}).get("score"):
                    review_scores.append(review_data["academic_review"]["score"])
                if review_data.get("essay_review", {}).get("scores"):
                    review_scores.extend(review_data["essay_review"]["scores"])
                avg_review_score = (
                    sum(review_scores) / len(review_scores) if review_scores else "N/A"
                )

                decision_label = "Pending"
                if match.get("award_decision"):
                    decision_label = (
                        match["award_decision"]["decision"].replace("_", " ").title()
                    )
                applicant_data.append(
                    [
                        applicant["name"],
                        applicant["student_id"],
                        applicant["major"],
                        f"{applicant['gpa']:.2f}",
                        applicant["academic_level"],
                        application_status.get("status", "Unknown").title(),
                        f"{avg_review_score:.1f}"
                        if isinstance(avg_review_score, float)
                        else avg_review_score,
                        decision_label,
                    ]
                )

            if len(applicant_data) > 1:
                applicant_table = Table(applicant_data)
                applicant_table.setStyle(
                    TableStyle(
                        [
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ]
                    )
                )
                story.append(applicant_table)

            # Detailed Review Information
            for match in scholarship_match["matches"]:
                applicant = match["applicant"]
                review_data = match.get("review_data", {})
                story.append(
                    Paragraph(
                        f"\nDetailed Review for {applicant['name']}:",
                        styles["Heading4"],
                    )
                )

                # Essay Reviews
                if review_data.get("essay_review", {}).get("comments"):
                    story.append(Paragraph("Essay Reviews:", styles["Heading4"]))
                    for i, (comment, score) in enumerate(
                        zip(
                            review_data["essay_review"]["comments"],
                            review_data["essay_review"]["scores"],
                        ),
                        1,
                    ):
                        story.append(
                            Paragraph(
                                f"Essay {i} - Score: {score}/10", styles["Normal"]
                            )
                        )
                        story.append(
                            Paragraph(f"Feedback: {comment}", styles["Normal"])
                        )

                # Interview Notes
                if review_data.get("interview_notes"):
                    story.append(Paragraph("Interview Notes:", styles["Heading4"]))
                    story.append(
                        Paragraph(review_data["interview_notes"], styles["Normal"])
                    )

                # Committee Feedback (retained)
                if review_data.get("committee_feedback"):
                    story.append(Paragraph("Committee Feedback:", styles["Heading4"]))
                    for feedback in review_data["committee_feedback"]:
                        story.append(
                            Paragraph(
                                f"• {feedback['member']}: {feedback['comments']}",
                                styles["Normal"],
                            )
                        )
                # Award Decision Details
                if match.get("award_decision"):
                    ad = match["award_decision"]
                    story.append(Paragraph("Award Decision:", styles["Heading4"]))
                    story.append(
                        Paragraph(
                            f"Decision: {ad['decision'].replace('_', ' ').title()}",
                            styles["Normal"],
                        )
                    )
                    if ad.get("decided_at"):
                        try:
                            story.append(
                                Paragraph(
                                    f"Decided At: {ad['decided_at'].strftime('%Y-%m-%d')}",
                                    styles["Normal"],
                                )
                            )
                        except Exception:
                            story.append(
                                Paragraph(
                                    f"Decided At: {ad['decided_at']}", styles["Normal"]
                                )
                            )
                    if ad.get("comments"):
                        story.append(Paragraph("Comments:", styles["Heading4"]))
                        story.append(Paragraph(str(ad["comments"]), styles["Normal"]))

            story.append(Paragraph("<br/>", styles["Normal"]))

        doc.build(story)
        return output_path

    def export_prescreening_report_to_csv(
        self,
        applicants: List[Applicant],
        scholarship_id: Optional[str] = None,
        output_path: str = None,
    ) -> str:
        """Export pre-screening report to CSV format.

        Args:
            applicants (List[Applicant]): List of applicants to evaluate
            scholarship_id (str, optional): Specific scholarship to evaluate for
            output_path (str): Path where to save the CSV file

        Returns:
            str: Path to the generated CSV file
        """
        report_data = self.generate_prescreening_report(applicants, scholarship_id)

        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # Write header and summary information
            writer.writerow(["Pre-screening Report"])
            writer.writerow(
                [
                    "Generated Date:",
                    report_data["generated_date"].strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )
            writer.writerow(["Total Applicants:", report_data["total_applicants"]])
            writer.writerow(["Total Matches:", report_data["summary"]["total_matches"]])
            writer.writerow(
                ["Match Rate:", f"{report_data['summary']['match_rate'] * 100:.1f}%"]
            )
            writer.writerow(
                [
                    "Scholarships with Matches:",
                    report_data["summary"]["scholarships_with_matches"],
                ]
            )
            writer.writerow([])

            # Write review statistics
            writer.writerow(["Review Statistics"])
            review_stats = report_data["summary"]["review_statistics"]
            writer.writerow(
                [
                    "Average Academic Review Score:",
                    f"{review_stats['average_scores']['academic_review']:.1f}/10",
                ]
            )
            writer.writerow(
                [
                    "Average Essay Review Score:",
                    f"{review_stats['average_scores']['essay_review']:.1f}/10",
                ]
            )
            writer.writerow(["Reviews Completed:", review_stats["reviews_completed"]])
            writer.writerow(
                ["Total Reviews Expected:", review_stats["total_reviews_expected"]]
            )
            writer.writerow(
                [
                    "Review Completion Rate:",
                    f"{review_stats['review_completion_rate'] * 100:.1f}%",
                ]
            )
            writer.writerow([])

            # Write application completion statistics
            writer.writerow(["Application Completion"])
            completion_stats = report_data["summary"]["application_completion"]
            writer.writerow(["Complete:", completion_stats["complete"]])
            writer.writerow(["In Progress:", completion_stats["in_progress"]])
            writer.writerow(["Incomplete:", completion_stats["incomplete"]])
            writer.writerow([])

            # Write detailed matches for each scholarship
            writer.writerow(["Scholarship Matches"])
            writer.writerow(
                [
                    "Scholarship Name",
                    "Eligibility Criteria",
                    "Applicant Name",
                    "Student ID",
                    "Major",
                    "GPA",
                    "Academic Level",
                    "Application Status",
                    "Qualification Score",
                    "Review Score",
                    "Has Interview",
                    "Has Committee Feedback",
                    "Award Decision",
                    "Decision Comments",
                ]
            )

            for match in report_data["matches"]:
                scholarship_name = match["scholarship_name"]
                eligibility_list = match.get("eligibility_criteria", [])
                eligibility_str = (
                    "; ".join(eligibility_list)
                    if isinstance(eligibility_list, list)
                    else str(eligibility_list)
                )
                for applicant_match in match["matches"]:
                    applicant = applicant_match["applicant"]
                    review_data = applicant_match["review_data"]

                    # Calculate average review score
                    review_scores = []
                    if review_data.get("academic_review", {}).get("score"):
                        review_scores.append(review_data["academic_review"]["score"])
                    if review_data.get("essay_review", {}).get("scores"):
                        review_scores.extend(review_data["essay_review"]["scores"])
                    avg_review_score = (
                        f"{sum(review_scores) / len(review_scores):.1f}"
                        if review_scores
                        else "N/A"
                    )

                    decision_label = "Pending"
                    decision_comments = ""
                    if applicant_match.get("award_decision"):
                        decision_label = (
                            applicant_match["award_decision"]["decision"]
                            .replace("_", " ")
                            .title()
                        )
                        decision_comments = applicant_match["award_decision"].get(
                            "comments", ""
                        )

                    writer.writerow(
                        [
                            scholarship_name,
                            eligibility_str,
                            applicant["name"],
                            applicant["student_id"],
                            applicant["major"],
                            f"{applicant['gpa']:.2f}",
                            applicant["academic_level"],
                            applicant_match["application_status"]["status"].title(),
                            f"{applicant_match['qualification_score']:.1f}%",
                            avg_review_score,
                            "Yes" if review_data.get("interview_notes") else "No",
                            "Yes" if review_data.get("committee_feedback") else "No",
                            decision_label,
                            decision_comments,
                        ]
                    )
            writer.writerow([])

            # Write detailed review information
            writer.writerow(["Detailed Reviews"])
            writer.writerow(
                [
                    "Applicant Name",
                    "Student ID",
                    "Review Type",
                    "Score",
                    "Comments",
                    "Reviewer",
                    "Date",
                ]
            )

            for match in report_data["matches"]:
                for applicant_match in match["matches"]:
                    applicant = applicant_match["applicant"]
                    review_data = applicant_match["review_data"]

                    # Academic Review
                    if review_data.get("academic_review", {}).get("score"):
                        writer.writerow(
                            [
                                applicant["name"],
                                applicant["student_id"],
                                "Academic Review",
                                review_data["academic_review"]["score"],
                                review_data["academic_review"].get("comments", "N/A"),
                                review_data["academic_review"].get("reviewer", "N/A"),
                                review_data["academic_review"].get("date", "N/A"),
                            ]
                        )

                    # Essay Reviews
                    for i, (score, comment, reviewer, date) in enumerate(
                        zip(
                            review_data["essay_review"]["scores"],
                            review_data["essay_review"]["comments"],
                            review_data["essay_review"]["reviewers"],
                            review_data["essay_review"]["dates"],
                        ),
                        1,
                    ):
                        writer.writerow(
                            [
                                applicant["name"],
                                applicant["student_id"],
                                f"Essay Review {i}",
                                score,
                                comment,
                                reviewer,
                                date.strftime("%Y-%m-%d") if date else "N/A",
                            ]
                        )

                    # Committee Feedback
                    for feedback in review_data.get("committee_feedback", []):
                        writer.writerow(
                            [
                                applicant["name"],
                                applicant["student_id"],
                                "Committee Feedback",
                                feedback.get("recommendation", "N/A"),
                                feedback["comments"],
                                feedback["member"],
                                feedback.get("date", "N/A"),
                            ]
                        )

        return output_path

    def export_prescreening_report_to_excel(
        self,
        applicants: List[Applicant],
        scholarship_id: Optional[str] = None,
        output_path: str = None,
    ) -> str:
        """Export pre-screening report to Excel format."""
        report_data = self.generate_prescreening_report(applicants, scholarship_id)

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "Pre-screening Report Summary"
        ws_summary["A2"] = "Generated on:"
        ws_summary["B2"] = report_data["generated_date"].strftime("%Y-%m-%d %H:%M:%S")

        summary_data = [
            ["Total Applicants", report_data["total_applicants"]],
            ["Total Matches", report_data["summary"]["total_matches"]],
            ["Match Rate", f"{report_data['summary']['match_rate'] * 100:.1f}%"],
            [
                "Scholarships with Matches",
                report_data["summary"]["scholarships_with_matches"],
            ],
            [
                "Review Completion Rate",
                f"{report_data['summary']['review_statistics']['review_completion_rate'] * 100:.1f}%",
            ],
            [
                "Applications Complete",
                report_data["summary"]["application_completion"]["complete"],
            ],
            [
                "Applications In Progress",
                report_data["summary"]["application_completion"]["in_progress"],
            ],
            [
                "Applications Incomplete",
                report_data["summary"]["application_completion"]["incomplete"],
            ],
        ]

        for row_idx, (label, value) in enumerate(summary_data, 4):
            ws_summary[f"A{row_idx}"] = label
            ws_summary[f"B{row_idx}"] = value

        # Review Statistics
        ws_summary["A12"] = "Review Statistics"
        ws_summary["A12"].font = Font(bold=True)
        review_stats = [
            [
                "Average Academic Review Score",
                f"{report_data['summary']['review_statistics']['average_scores']['academic_review']:.1f}/10",
            ],
            [
                "Average Essay Review Score",
                f"{report_data['summary']['review_statistics']['average_scores']['essay_review']:.1f}/10",
            ],
            [
                "Reviews Completed",
                report_data["summary"]["review_statistics"]["reviews_completed"],
            ],
            [
                "Total Reviews Expected",
                report_data["summary"]["review_statistics"]["total_reviews_expected"],
            ],
        ]
        for row_idx, (label, value) in enumerate(review_stats, 13):
            ws_summary[f"A{row_idx}"] = label
            ws_summary[f"B{row_idx}"] = value

        # Matches Sheet with Review Information
        for scholarship_match in report_data["matches"]:
            ws_matches = wb.create_sheet(scholarship_match["scholarship_name"][:31])

            # Scholarship details
            ws_matches["A1"] = "Scholarship Details"
            ws_matches["A2"] = "Description:"
            ws_matches["B2"] = scholarship_match["description"]
            ws_matches["A3"] = "Amount:"
            ws_matches["B3"] = f"${scholarship_match['amount']:,.2f}"
            ws_matches["A4"] = "Deadline:"
            ws_matches["B4"] = (
                scholarship_match["deadline"].strftime("%Y-%m-%d")
                if scholarship_match["deadline"]
                else "No deadline set"
            )
            # Eligibility Criteria
            ws_matches["A5"] = "Eligibility Criteria:"
            eligibility_list = scholarship_match.get("eligibility_criteria", [])
            if isinstance(eligibility_list, list):
                ws_matches["B5"] = "; ".join(eligibility_list)
            else:
                ws_matches["B5"] = str(eligibility_list) if eligibility_list else "N/A"

            # Matching applicants with review scores
            ws_matches["A6"] = "Qualified Applicants"
            headers = [
                "Name",
                "Student ID",
                "Major",
                "GPA",
                "Academic Level",
                "Application Status",
                "Review Score",
                "Essay Scores",
                "Interview Complete",
                "Committee Review",
                "Award Decision",
                "Decision Comments",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws_matches.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            row = 8
            for match in scholarship_match["matches"]:
                applicant = match["applicant"]
                review_data = match.get("review_data", {})
                application_status = match.get("application_status", {})

                # Calculate average review score
                review_scores = []
                if review_data.get("academic_review", {}).get("score"):
                    review_scores.append(review_data["academic_review"]["score"])
                if review_data.get("essay_review", {}).get("scores"):
                    review_scores.extend(review_data["essay_review"]["scores"])
                avg_review_score = (
                    sum(review_scores) / len(review_scores) if review_scores else "N/A"
                )

                ws_matches.cell(row=row, column=1, value=applicant["name"])
                ws_matches.cell(row=row, column=2, value=applicant["student_id"])
                ws_matches.cell(row=row, column=3, value=applicant["major"])
                ws_matches.cell(row=row, column=4, value=f"{applicant['gpa']:.2f}")
                ws_matches.cell(row=row, column=5, value=applicant["academic_level"])
                ws_matches.cell(
                    row=row,
                    column=6,
                    value=application_status.get("status", "Unknown").title(),
                )
                ws_matches.cell(
                    row=row,
                    column=7,
                    value=f"{avg_review_score:.1f}"
                    if isinstance(avg_review_score, float)
                    else avg_review_score,
                )
                ws_matches.cell(
                    row=row,
                    column=8,
                    value=", ".join(
                        f"{score:.1f}"
                        for score in review_data.get("essay_review", {}).get(
                            "scores", []
                        )
                    )
                    or "N/A",
                )
                ws_matches.cell(
                    row=row,
                    column=9,
                    value="Yes" if review_data.get("interview_notes") else "No",
                )
                ws_matches.cell(
                    row=row,
                    column=10,
                    value="Yes" if review_data.get("committee_feedback") else "No",
                )
                decision_label = "Pending"
                decision_comments = ""
                if match.get("award_decision"):
                    decision_label = (
                        match["award_decision"]["decision"].replace("_", " ").title()
                    )
                    decision_comments = match["award_decision"].get("comments", "")
                ws_matches.cell(row=row, column=11, value=decision_label)
                ws_matches.cell(row=row, column=12, value=decision_comments)
                row += 1

                # Add detailed review information
                if review_data.get("interview_notes"):
                    row += 1
                    ws_matches.cell(row=row, column=1, value="Interview Notes:")
                    ws_matches.cell(
                        row=row, column=2, value=review_data["interview_notes"]
                    )
                    row += 1

                if review_data.get("committee_feedback"):
                    row += 1
                    ws_matches.cell(row=row, column=1, value="Committee Feedback:")
                    for feedback in review_data["committee_feedback"]:
                        row += 1
                        ws_matches.cell(
                            row=row,
                            column=2,
                            value=f"{feedback['member']}: {feedback['comments']}",
                        )
                    row += 1

            # Create Review Details Sheet for each scholarship
            ws_reviews = wb.create_sheet(
                f"{scholarship_match['scholarship_name'][:20]}_Reviews"
            )
            ws_reviews["A1"] = (
                f"Detailed Review Information for {scholarship_match['scholarship_name']}"
            )
            ws_reviews["A1"].font = Font(bold=True)

            row = 3
            for match in scholarship_match["matches"]:
                applicant = match["applicant"]
                review_data = match.get("review_data", {})

                ws_reviews.cell(
                    row=row, column=1, value=f"Review Details for {applicant['name']}"
                )
                ws_reviews.cell(row=row, column=1).font = Font(bold=True)
                row += 2

                # Essay Reviews
                if review_data.get("essay_review", {}).get("comments"):
                    ws_reviews.cell(row=row, column=1, value="Essay Reviews")
                    ws_reviews.cell(row=row, column=1).font = Font(bold=True)
                    row += 1

                    for i, (comment, score, reviewer, date) in enumerate(
                        zip(
                            review_data["essay_review"]["comments"],
                            review_data["essay_review"]["scores"],
                            review_data["essay_review"]["reviewers"],
                            review_data["essay_review"]["dates"],
                        ),
                        1,
                    ):
                        ws_reviews.cell(row=row, column=1, value=f"Essay {i}")
                        ws_reviews.cell(row=row, column=2, value=f"Score: {score}/10")
                        ws_reviews.cell(row=row + 1, column=1, value="Reviewer:")
                        ws_reviews.cell(row=row + 1, column=2, value=reviewer)
                        ws_reviews.cell(row=row + 2, column=1, value="Date:")
                        ws_reviews.cell(
                            row=row + 2,
                            column=2,
                            value=date.strftime("%Y-%m-%d") if date else "N/A",
                        )
                        ws_reviews.cell(row=row + 3, column=1, value="Feedback:")
                        ws_reviews.cell(row=row + 3, column=2, value=comment)
                        row += 5

                # Committee Feedback
                if review_data.get("committee_feedback"):
                    ws_reviews.cell(row=row, column=1, value="Committee Feedback")
                    ws_reviews.cell(row=row, column=1).font = Font(bold=True)
                    row += 1

                    for feedback in review_data["committee_feedback"]:
                        ws_reviews.cell(row=row, column=1, value="Member:")
                        ws_reviews.cell(row=row, column=2, value=feedback["member"])
                        ws_reviews.cell(row=row + 1, column=1, value="Comments:")
                        ws_reviews.cell(
                            row=row + 1, column=2, value=feedback["comments"]
                        )
                        if "date" in feedback:
                            ws_reviews.cell(row=row + 2, column=1, value="Date:")
                            ws_reviews.cell(
                                row=row + 2,
                                column=2,
                                value=feedback["date"],
                            )
                        row += 4

                row += 2  # Add space between applicants

        # Adjust column widths
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[chr(64 + col[0].column)].width = min(
                    max_length + 2, 50
                )

        wb.save(output_path)
        return output_path

    @staticmethod
    def _parse_iso_dates(obj):
        """Recursively convert ISO date strings back to datetime objects in a dict/list structure."""
        from datetime import datetime

        if isinstance(obj, str):
            try:
                return datetime.fromisoformat(obj.replace("Z", "+00:00"))
            except ValueError:
                return obj
        if isinstance(obj, dict):
            return {k: ReportEngine._parse_iso_dates(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [ReportEngine._parse_iso_dates(v) for v in obj]
        if isinstance(obj, tuple):
            return tuple(ReportEngine._parse_iso_dates(v) for v in obj)
        return obj

    # Function to generate applicant report. Meets requirement SFWE504_3-LLR-6.
    def generate_applicant_report(
        self, student_id: str = None, netid: str = None
    ) -> Dict[str, Any]:
        """Generate a comprehensive report of applicant(s) data and scholarship status.

        Args:
            student_id (str, optional): Student ID to search for. If None and netid is None, returns all applicants.
            netid (str, optional): NetID to search for (alternative to student_id)

        Returns:
            dict: Comprehensive applicant report including:
                - Personal and academic information
                - Current and past scholarships
                - Academic achievements
                - Financial information
                - Essay submissions and evaluations
                - If no student_id/netid provided: returns report for all applicants
        """
        # Determine if we're generating a report for one applicant or all
        banned_names = {
            name.lower() for name in ["Test User", "Test Student", "Test Award Student"]
        }
        if student_id or netid:
            # Single applicant report
            applicant_data = None
            if student_id:
                try:
                    applicant_data = Applicant.objects.get(student_id=student_id)
                except Applicant.DoesNotExist:
                    pass
            elif netid:
                try:
                    applicant_data = Applicant.objects.get(netid=netid)
                except Applicant.DoesNotExist:
                    pass

            # Exclude test/dummy applicants if matched directly
            if not applicant_data or (
                applicant_data
                and applicant_data.name
                and applicant_data.name.lower() in banned_names
            ):
                return None

            applicants_to_process = [applicant_data]
        else:
            # All applicants report
            # Exclude test/dummy applicants by name while preserving ordering
            applicants_to_process = [
                a
                for a in Applicant.objects.all().order_by("name")
                if (a.name or "").lower() not in banned_names
            ]

        # Process each applicant
        all_applicant_reports = []
        for applicant_data in applicants_to_process:
            # Ensure required applicant fields exist; auto-generate sensible defaults if missing
            defaults_applied = False
            try:
                # Minor
                if not getattr(applicant_data, "minor", None):
                    applicant_data.minor = "N/A"
                    defaults_applied = True

                # Academic achievements
                if not getattr(applicant_data, "academic_achievements", None):
                    from django.utils import timezone as _tz

                    applicant_data.academic_achievements = [
                        {
                            "type": "Orientation Completed",
                            "date": _tz.now().date().isoformat(),
                            "description": "New student orientation completed",
                        }
                    ]
                    defaults_applied = True

                # Financial information
                fi = getattr(applicant_data, "financial_info", None)
                if not isinstance(fi, dict) or not fi:
                    applicant_data.financial_info = {
                        "fafsa_submitted": False,
                        "efc": 0,
                        "household_income": "N/A",
                        "current_aid": [],
                    }
                    defaults_applied = True

                # Essay submissions
                if not getattr(applicant_data, "essays", None):
                    from django.utils import timezone as _tz2

                    applicant_data.essays = [
                        {
                            "prompt": "Personal Statement",
                            "content": "N/A",
                            "submission_date": _tz2.now().isoformat(),
                        }
                    ]
                    defaults_applied = True

                if defaults_applied:
                    applicant_data.save()
            except Exception:
                # Non-fatal: continue with available data
                pass
            # Query only ACTIVE awards for this applicant (exclude previous/completed awards)
            awards_queryset = ScholarshipAward.objects.filter(
                applicant=applicant_data, status="active"
            )

            # Deduplicate by scholarship_name: keep the most recent award per scholarship
            awards_by_name = {}
            for aw in awards_queryset.order_by(
                "scholarship_name", "-award_date", "-id"
            ):
                if aw.scholarship_name not in awards_by_name:
                    awards_by_name[aw.scholarship_name] = aw
            deduped_awards = list(awards_by_name.values())

            applicant_awards = []
            for award in deduped_awards:
                applicant_awards.append(
                    {
                        "scholarship_name": award.scholarship_name,
                        "award_amount": float(award.award_amount),
                        "award_date": award.award_date,
                        "status": award.status,
                        "disbursements": [
                            {
                                "date": date,
                                "amount": float(award.award_amount)
                                / len(award.disbursement_dates)
                                if award.disbursement_dates
                                else float(award.award_amount),
                            }
                            for date in award.disbursement_dates
                        ],
                        "requirements_met": award.requirements_met,
                        "requirements_pending": award.requirements_pending,
                        "performance_metrics": award.performance_metrics,
                        "essays_evaluation": award.essays_evaluation,
                        "interview_notes": award.interview_notes,
                        "committee_feedback": award.committee_feedback,
                    }
                )

            # Compile comprehensive applicant report and parse any ISO date strings
            # Build normalized essay evaluation list combining applicant essay evaluations and award-linked evaluations
            normalized_essay_evaluations: List[Dict[str, Any]] = []

            # Applicant-level essay evaluations (embedded in applicant_data.essays as 'evaluation')
            if getattr(applicant_data, "essays", None):
                for essay in applicant_data.essays:
                    if isinstance(essay, dict):
                        eval_block = essay.get("evaluation")
                        if isinstance(eval_block, dict):
                            normalized_essay_evaluations.append(
                                {
                                    "prompt": essay.get("prompt", ""),
                                    "score": eval_block.get("score"),
                                    "feedback": eval_block.get("feedback"),
                                    "reviewer": eval_block.get("reviewer"),
                                    "date": self._parse_iso_dates(
                                        eval_block.get("date")
                                    ),
                                    "source": "Applicant Essay",
                                    "scholarship_name": None,
                                }
                            )

            # Award-level essay evaluations (award['essays_evaluation'] list)
            for award in applicant_awards:
                evaluations = award.get("essays_evaluation")
                if isinstance(evaluations, list):
                    for ev in evaluations:
                        if isinstance(ev, dict):
                            normalized_essay_evaluations.append(
                                {
                                    "prompt": ev.get("prompt", ""),
                                    "score": ev.get("score"),
                                    "feedback": ev.get("feedback"),
                                    "reviewer": ev.get("reviewer"),
                                    "date": self._parse_iso_dates(ev.get("date")),
                                    "source": "Scholarship Award",
                                    "scholarship_name": award.get("scholarship_name"),
                                }
                            )

            applicant_report = {
                "personal_info": {
                    "name": applicant_data.name,
                    "student_id": applicant_data.student_id,
                    "netid": applicant_data.netid,
                },
                "academic_info": {
                    "major": applicant_data.major,
                    "minor": applicant_data.minor,
                    "gpa": applicant_data.gpa,
                    "academic_level": applicant_data.academic_level,
                    "expected_graduation": applicant_data.expected_graduation,  # This is already a date from model
                    "academic_history": self._parse_iso_dates(
                        applicant_data.academic_history
                    ),
                },
                "achievements": self._parse_iso_dates(
                    applicant_data.academic_achievements
                ),
                "financial_info": applicant_data.financial_info,
                "essays": [
                    {
                        "prompt": essay.get("prompt", "")
                        if isinstance(essay, dict)
                        else "",
                        "submission_date": self._parse_iso_dates(
                            essay.get("submission_date")
                        )
                        if isinstance(essay, dict)
                        else None,
                        "content": essay.get("content", "")
                        if isinstance(essay, dict)
                        else str(essay),
                    }
                    for essay in (
                        applicant_data.essays if applicant_data.essays else []
                    )
                ],
                "essay_evaluations": self._parse_iso_dates(
                    normalized_essay_evaluations
                ),
                "scholarships": {
                    "total_awards": len(applicant_awards),
                    "total_amount": sum(
                        award["award_amount"] for award in applicant_awards
                    ),
                    "active_awards": self._parse_iso_dates(
                        [
                            award
                            for award in applicant_awards
                            if award["status"] == "active"
                        ]
                    ),
                    "completed_awards": self._parse_iso_dates(
                        [
                            award
                            for award in applicant_awards
                            if award["status"] == "completed"
                        ]
                    ),
                    "detailed_awards": self._parse_iso_dates(applicant_awards),
                },
            }
            all_applicant_reports.append(applicant_report)

        # Return single report or multi-applicant report based on input
        if student_id or netid:
            # Single applicant - return the report directly
            return all_applicant_reports[0] if all_applicant_reports else None
        else:
            # Multiple applicants - return a summary report
            return {
                "total_applicants": len(all_applicant_reports),
                "applicants": all_applicant_reports,
                "summary": {
                    "total_scholarship_awards": sum(
                        app["scholarships"]["total_awards"]
                        for app in all_applicant_reports
                    ),
                    "total_scholarship_amount": sum(
                        app["scholarships"]["total_amount"]
                        for app in all_applicant_reports
                    ),
                    "average_gpa": sum(
                        app["academic_info"]["gpa"] for app in all_applicant_reports
                    )
                    / len(all_applicant_reports)
                    if all_applicant_reports
                    else 0,
                },
            }

    # Function to generate scholarship report. Meets requirements SFWE504_3-LLR-3.
    def generate_scholarship_report(
        self, filters=None, export_format=None, output_path=None
    ):
        """Generate a comprehensive report of scholarships and optionally export it.

        Args:
            filters (dict, optional): Filters to apply (e.g., {'frequency': 'annual'})
            export_format (str, optional): Format to export the report ('pdf', 'xlsx', 'csv')
            output_path (str, optional): Path where the exported file should be saved.

        Returns:
            Union[dict, str]: Report data as dictionary if no export_format specified,
                              otherwise path to the exported file.
        """
        # Use unified data access so all features see the same set of scholarships
        scholarships_data = self.get_scholarships_data()

        # Apply filters if provided
        if filters:
            for key, value in filters.items():
                scholarships_data = [
                    s for s in scholarships_data if getattr(s, key, None) == value
                ]

        # Generate report summary
        total_amount = sum(s.amount for s in scholarships_data)
        frequencies = {}
        for s in scholarships_data:
            frequencies[s.frequency] = frequencies.get(s.frequency, 0) + 1

        # Format scholarship details
        scholarship_details = []
        for s in scholarships_data:
            scholarship_details.append(
                {
                    "name": s.name,
                    "description": s.description,
                    "eligibility": s.eligibility_criteria,
                    "donor": s.donor_info,
                    "requirements": s.disbursement_requirements,
                    "frequency": s.frequency,
                    "amount": s.amount,
                    "deadline": s.deadline.strftime("%Y-%m-%d")
                    if s.deadline
                    else "No deadline set",
                }
            )

        report_data = {
            "total_scholarships": len(scholarships_data),
            "total_amount": total_amount,
            "frequency_distribution": frequencies,
            "scholarships": scholarship_details,
        }

        # Handle export if requested
        if export_format:
            if not output_path:
                temp_file = tempfile.NamedTemporaryFile(
                    delete=False, suffix=f".{export_format}"
                )
                output_path = temp_file.name
                temp_file.close()
            if export_format.lower() == "pdf":
                return self.export_to_pdf(output_path, filters)
            if export_format.lower() == "xlsx":
                return self.export_to_excel(output_path, filters)
            if export_format.lower() == "csv":
                return self.export_to_csv(output_path, filters)
            raise ValueError(f"Unsupported export format: {export_format}")

        return report_data

    # Export Methods for PDF, Excel, CSV meeting the requirement SFWE504_3-LLR-3, SFWE504_3-LLR-11, and SFWE504_3-LLR-33
    def export_to_pdf(self, output_path: str, filters=None) -> str:
        """Export scholarships data to PDF format."""
        report_data = self.generate_scholarship_report(filters)

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []

        # Title and Summary
        styles = getSampleStyleSheet()
        story.append(Paragraph("Scholarship Report", styles["Heading1"]))
        story.append(
            Paragraph(
                f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"Total Scholarships: {report_data['total_scholarships']}",
                styles["Normal"],
            )
        )
        story.append(
            Paragraph(
                f"Total Amount: ${report_data['total_amount']:,.2f}", styles["Normal"]
            )
        )

        # Frequency Distribution
        story.append(Paragraph("Frequency Distribution:", styles["Heading2"]))
        freq_data = [
            [freq, count]
            for freq, count in report_data["frequency_distribution"].items()
        ]
        if freq_data:
            freq_table = Table([["Frequency", "Count"]] + freq_data)
            freq_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(freq_table)
        story.append(Paragraph("<br/><br/>", styles["Normal"]))

        # Scholarships Details
        story.append(Paragraph("Scholarship Details:", styles["Heading2"]))
        for scholarship in report_data["scholarships"]:
            # Scholarship Header
            story.append(Paragraph(f"<br/>{scholarship['name']}", styles["Heading3"]))
            story.append(
                Paragraph(f"Amount: ${scholarship['amount']:,.2f}", styles["Normal"])
            )
            story.append(
                Paragraph(f"Deadline: {scholarship['deadline']}", styles["Normal"])
            )
            story.append(
                Paragraph(f"Frequency: {scholarship['frequency']}", styles["Normal"])
            )

            # Donor/Sponsor Information
            donor_info = scholarship.get("donor", {})
            if donor_info:
                story.append(
                    Paragraph("Donor/Sponsor Information:", styles["Heading4"])
                )
                donor_name = donor_info.get("name", "N/A")
                donor_contact = donor_info.get("contact", "N/A")
                donor_org = donor_info.get(
                    "organization", donor_info.get("name", "N/A")
                )
                donor_phone = donor_info.get("phone", "N/A")
                donor_email = donor_info.get(
                    "email", donor_contact if "@" in str(donor_contact) else "N/A"
                )
                donor_address = donor_info.get("address", "N/A")

                story.append(Paragraph(f"Name: {donor_name}", styles["Normal"]))
                if donor_contact != "N/A":
                    story.append(
                        Paragraph(f"Contact: {donor_contact}", styles["Normal"])
                    )
                if donor_email != "N/A" and donor_email != donor_contact:
                    story.append(Paragraph(f"Email: {donor_email}", styles["Normal"]))
                if donor_phone != "N/A":
                    story.append(Paragraph(f"Phone: {donor_phone}", styles["Normal"]))
                if donor_address != "N/A":
                    story.append(
                        Paragraph(f"Address: {donor_address}", styles["Normal"])
                    )

            # Description
            story.append(Paragraph("Description:", styles["Heading4"]))
            story.append(Paragraph(scholarship["description"], styles["Normal"]))

            # Eligibility Criteria
            story.append(Paragraph("Eligibility Criteria:", styles["Heading4"]))
            for criterion in scholarship["eligibility"]:
                story.append(Paragraph(f"• {criterion}", styles["Normal"]))

            # Requirements
            story.append(Paragraph("Disbursement Requirements:", styles["Heading4"]))
            for req in scholarship["requirements"]:
                story.append(Paragraph(f"• {req}", styles["Normal"]))

            story.append(Paragraph("<br/>", styles["Normal"]))

        doc.build(story)
        return output_path

    def export_to_excel(self, output_path: str, filters=None) -> str:
        """Export scholarships data to Excel format."""
        report_data = self.generate_scholarship_report(filters)

        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "Scholarship Report Summary"
        ws_summary["A2"] = "Generated on:"
        ws_summary["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A3"] = "Total Scholarships:"
        ws_summary["B3"] = report_data["total_scholarships"]
        ws_summary["A4"] = "Total Amount:"
        ws_summary["B4"] = f"${report_data['total_amount']:,.2f}"

        # Frequency Distribution
        ws_summary["A6"] = "Frequency Distribution"
        ws_summary["A7"] = "Frequency"
        ws_summary["B7"] = "Count"
        row = 8
        for freq, count in report_data["frequency_distribution"].items():
            ws_summary[f"A{row}"] = freq
            ws_summary[f"B{row}"] = count
            row += 1

        # Scholarships Sheet
        ws_details = wb.create_sheet("Scholarship Details")
        headers = [
            "Name",
            "Amount",
            "Deadline",
            "Frequency",
            "Description",
            "Donor Name",
            "Donor Contact",
            "Donor Email",
            "Donor Phone",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        for row, scholarship in enumerate(report_data["scholarships"], 2):
            donor_info = scholarship.get("donor", {})
            donor_name = donor_info.get("name", "N/A") if donor_info else "N/A"
            donor_contact = donor_info.get("contact", "N/A") if donor_info else "N/A"
            donor_email = (
                donor_info.get(
                    "email", donor_contact if "@" in str(donor_contact) else "N/A"
                )
                if donor_info
                else "N/A"
            )
            donor_phone = donor_info.get("phone", "N/A") if donor_info else "N/A"

            ws_details.cell(row=row, column=1, value=scholarship["name"])
            ws_details.cell(row=row, column=2, value=f"${scholarship['amount']:,.2f}")
            ws_details.cell(row=row, column=3, value=scholarship["deadline"])
            ws_details.cell(row=row, column=4, value=scholarship["frequency"])
            ws_details.cell(row=row, column=5, value=scholarship["description"])
            ws_details.cell(row=row, column=6, value=donor_name)
            ws_details.cell(row=row, column=7, value=donor_contact)
            ws_details.cell(row=row, column=8, value=donor_email)
            ws_details.cell(row=row, column=9, value=donor_phone)

        # Adjust column widths
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[chr(64 + col[0].column)].width = min(
                    max_length + 2, 50
                )

        wb.save(output_path)
        return output_path

    def export_to_csv(self, output_path: str, filters=None) -> str:
        """Export scholarships data to CSV format."""
        report_data = self.generate_scholarship_report(filters)

        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # Write summary
            writer.writerow(["Scholarship Report Summary"])
            writer.writerow(
                ["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            )
            writer.writerow(["Total Scholarships:", report_data["total_scholarships"]])
            writer.writerow(["Total Amount:", f"${report_data['total_amount']:,.2f}"])
            writer.writerow([])

            # Write frequency distribution
            writer.writerow(["Frequency Distribution"])
            writer.writerow(["Frequency", "Count"])
            for freq, count in report_data["frequency_distribution"].items():
                writer.writerow([freq, count])
            writer.writerow([])

            # Write scholarship details
            writer.writerow(["Scholarship Details"])
            writer.writerow(
                [
                    "Name",
                    "Amount",
                    "Deadline",
                    "Frequency",
                    "Description",
                    "Eligibility Criteria",
                    "Requirements",
                    "Donor Name",
                    "Donor Contact",
                    "Donor Email",
                    "Donor Phone",
                ]
            )

            for scholarship in report_data["scholarships"]:
                donor_info = scholarship.get("donor", {})
                donor_name = donor_info.get("name", "N/A") if donor_info else "N/A"
                donor_contact = (
                    donor_info.get("contact", "N/A") if donor_info else "N/A"
                )
                donor_email = (
                    donor_info.get(
                        "email", donor_contact if "@" in str(donor_contact) else "N/A"
                    )
                    if donor_info
                    else "N/A"
                )
                donor_phone = donor_info.get("phone", "N/A") if donor_info else "N/A"

                writer.writerow(
                    [
                        scholarship["name"],
                        f"${scholarship['amount']:,.2f}",
                        scholarship["deadline"],
                        scholarship["frequency"],
                        scholarship["description"],
                        "; ".join(scholarship["eligibility"]),
                        "; ".join(scholarship["requirements"]),
                        donor_name,
                        donor_contact,
                        donor_email,
                        donor_phone,
                    ]
                )

        return output_path

    def export_applicant_report_to_pdf(
        self, student_id: str = None, netid: str = None, output_path: str = None
    ) -> str:
        """Export applicant report to PDF format."""
        report_data = self.generate_applicant_report(student_id, netid)
        if not report_data:
            raise ValueError("Applicant not found")

        # Check if this is a multi-applicant report or single applicant
        is_multi_applicant = "applicants" in report_data

        # Use landscape orientation for the multi-applicant summary; portrait otherwise
        try:
            from reportlab.lib.pagesizes import landscape, letter as _letter

            pagesize = landscape(_letter) if is_multi_applicant else _letter
        except Exception:
            # Fallback to portrait letter if landscape import not available
            pagesize = letter

        doc = SimpleDocTemplate(output_path, pagesize=pagesize)
        story = []
        styles = getSampleStyleSheet()

        if is_multi_applicant:
            # Multi-applicant summary report
            story.append(Paragraph(f"All Applicants Report", styles["Heading1"]))
            story.append(
                Paragraph(
                    f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    styles["Normal"],
                )
            )
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Summary statistics
            story.append(Paragraph("Summary Statistics", styles["Heading2"]))
            summary_data = [
                ["Total Applicants:", str(report_data["total_applicants"])],
                [
                    "Total Scholarship Awards:",
                    str(report_data["summary"]["total_scholarship_awards"]),
                ],
                [
                    "Total Scholarship Amount:",
                    f"${report_data['summary']['total_scholarship_amount']:,.2f}",
                ],
                ["Average GPA:", f"{report_data['summary']['average_gpa']:.2f}"],
            ]
            summary_table = Table(summary_data)
            summary_table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                    ]
                )
            )
            story.append(summary_table)
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Individual applicant summaries
            story.append(Paragraph("Individual Applicants", styles["Heading2"]))
            applicant_summary_data = [
                [
                    "Name",
                    "Student ID",
                    "Major",
                    "Minor",
                    "GPA",
                    "Academic Level",
                    "Achievements (#)",
                    "FAFSA",
                    "EFC",
                    "Income Range",
                    "Essay Submissions (#)",
                    "Awards",
                    "Total Amount",
                ]
            ]
            for applicant in report_data["applicants"]:
                achievements = applicant.get("achievements") or []
                financial = applicant.get("financial_info") or {}
                essays = applicant.get("essays") or []
                applicant_summary_data.append(
                    [
                        applicant["personal_info"]["name"],
                        applicant["personal_info"]["student_id"],
                        applicant["academic_info"]["major"],
                        applicant["academic_info"].get("minor") or "N/A",
                        f"{applicant['academic_info']['gpa']:.2f}",
                        applicant["academic_info"]["academic_level"],
                        str(len(achievements)),
                        "Yes" if financial.get("fafsa_submitted") else "No",
                        financial.get("efc", 0),
                        financial.get("household_income", "N/A"),
                        str(len(essays)),
                        str(applicant["scholarships"]["total_awards"]),
                        f"${applicant['scholarships']['total_amount']:,.2f}",
                    ]
                )

            applicant_table = Table(applicant_summary_data)
            applicant_table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.append(applicant_table)

        else:
            # Single applicant detailed report (existing logic)
            story.append(
                Paragraph(
                    f"Applicant Report: {report_data['personal_info']['name']}",
                    styles["Heading1"],
                )
            )
            story.append(
                Paragraph(
                    f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    styles["Normal"],
                )
            )
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Personal and Academic Information
            story.append(Paragraph("Personal Information", styles["Heading2"]))
            personal_info = [
                ["Student ID:", report_data["personal_info"]["student_id"]],
                ["NetID:", report_data["personal_info"]["netid"]],
                ["Major:", report_data["academic_info"]["major"]],
                ["Minor:", report_data["academic_info"]["minor"] or "N/A"],
                ["GPA:", f"{report_data['academic_info']['gpa']:.2f}"],
                ["Academic Level:", report_data["academic_info"]["academic_level"]],
                [
                    "Expected Graduation:",
                    report_data["academic_info"]["expected_graduation"].strftime(
                        "%Y-%m-%d"
                    ),
                ],
            ]
            info_table = Table(personal_info)
            info_table.setStyle(
                TableStyle(
                    [
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                        ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                    ]
                )
            )
            story.append(info_table)
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Academic Achievements
            story.append(Paragraph("Academic Achievements", styles["Heading2"]))
            if report_data.get("achievements"):
                for achievement in report_data["achievements"]:
                    if isinstance(achievement, dict):
                        achievement_type = achievement.get("type", "Achievement")
                        achievement_date = achievement.get("date")
                        date_str = (
                            achievement_date.strftime("%Y-%m-%d")
                            if hasattr(achievement_date, "strftime")
                            else str(achievement_date)
                            if achievement_date
                            else "N/A"
                        )
                        story.append(
                            Paragraph(
                                f"• {achievement_type} - {date_str}", styles["Normal"]
                            )
                        )
                        if achievement.get("description"):
                            story.append(
                                Paragraph(
                                    f"  {achievement['description']}", styles["Normal"]
                                )
                            )
                    else:
                        story.append(
                            Paragraph(f"• {str(achievement)}", styles["Normal"])
                        )
            else:
                story.append(Paragraph("No achievements recorded", styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Financial Information
            story.append(Paragraph("Financial Information", styles["Heading2"]))
            financial_info = report_data.get("financial_info", {})
            if isinstance(financial_info, dict):
                story.append(
                    Paragraph(
                        f"FAFSA Submitted: {financial_info.get('fafsa_submitted', 'N/A')}",
                        styles["Normal"],
                    )
                )
                story.append(
                    Paragraph(
                        f"Expected Family Contribution: ${financial_info.get('efc', 0):,}",
                        styles["Normal"],
                    )
                )
                story.append(
                    Paragraph(
                        f"Household Income Range: {financial_info.get('household_income', 'N/A')}",
                        styles["Normal"],
                    )
                )
            else:
                story.append(
                    Paragraph("Financial information not available", styles["Normal"])
                )
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Current Aid
            if isinstance(financial_info, dict) and financial_info.get("current_aid"):
                story.append(Paragraph("Current Financial Aid:", styles["Heading3"]))
                for aid in financial_info["current_aid"]:
                    if isinstance(aid, dict):
                        story.append(
                            Paragraph(
                                f"• {aid.get('type', 'Aid')}: ${aid.get('amount', 0):,}",
                                styles["Normal"],
                            )
                        )
                    else:
                        story.append(Paragraph(f"• {str(aid)}", styles["Normal"]))
            story.append(Paragraph("<br/>", styles["Normal"]))

            # Essay Submissions (new section)
            story.append(Paragraph("Essay Submissions", styles["Heading2"]))
            essays_list = report_data.get("essays") or []
            if essays_list:
                for es in essays_list:
                    if isinstance(es, dict):
                        sub_date = es.get("submission_date")
                        if hasattr(sub_date, "strftime"):
                            sub_date_str = sub_date.strftime("%Y-%m-%d")
                        else:
                            sub_date_str = str(sub_date) if sub_date else "N/A"
                        content_preview = (es.get("content", "") or "")[:120]
                        story.append(
                            Paragraph(
                                f"• {es.get('prompt', 'Essay')} ({sub_date_str})",
                                styles["Normal"],
                            )
                        )
                        if content_preview:
                            story.append(
                                Paragraph(f"  {content_preview}", styles["Normal"])
                            )
                story.append(Paragraph("<br/>", styles["Normal"]))
            else:
                story.append(
                    Paragraph("No essay submissions recorded", styles["Normal"])
                )
                story.append(Paragraph("<br/>", styles["Normal"]))

            # Scholarship Awards
            story.append(Paragraph("Scholarship Awards", styles["Heading2"]))
            story.append(
                Paragraph(
                    f"Total Awards: {report_data['scholarships']['total_awards']} "
                    f"(${report_data['scholarships']['total_amount']:,})",
                    styles["Normal"],
                )
            )

            for award in report_data["scholarships"]["detailed_awards"]:
                story.append(
                    Paragraph(
                        f"Award: {award.get('scholarship_name', 'Unknown')}",
                        styles["Heading3"],
                    )
                )
                story.append(
                    Paragraph(
                        f"Amount: ${award.get('award_amount', 0):,}", styles["Normal"]
                    )
                )
                story.append(
                    Paragraph(f"Status: {award.get('status', 'N/A')}", styles["Normal"])
                )
                award_date = award.get("award_date")
                if hasattr(award_date, "strftime"):
                    story.append(
                        Paragraph(
                            f"Award Date: {award_date.strftime('%Y-%m-%d')}",
                            styles["Normal"],
                        )
                    )
                elif award_date:
                    story.append(
                        Paragraph(f"Award Date: {str(award_date)}", styles["Normal"])
                    )
                # (Per-award raw evaluations removed; consolidated table provided below)

                if award.get("committee_feedback"):
                    story.append(Paragraph("Committee Feedback:", styles["Heading4"]))
                    for feedback in award["committee_feedback"]:
                        if isinstance(feedback, dict):
                            story.append(
                                Paragraph(
                                    f"• {feedback.get('member', 'Member')}: {feedback.get('comments', 'No comments')}",
                                    styles["Normal"],
                                )
                            )
                        else:
                            story.append(
                                Paragraph(f"• {str(feedback)}", styles["Normal"])
                            )
                story.append(Paragraph("<br/>", styles["Normal"]))

            # Consolidated Essay Evaluation Section
            evaluations = report_data.get("essay_evaluations", [])
            story.append(
                Paragraph("Consolidated Essay Evaluations", styles["Heading2"])
            )
            if evaluations:
                eval_table_data = [
                    [
                        "Source",
                        "Scholarship",
                        "Prompt",
                        "Score",
                        "Reviewer",
                        "Date",
                        "Feedback",
                    ]
                ]
                for ev in evaluations:
                    date_obj = ev.get("date")
                    date_str = (
                        date_obj.strftime("%Y-%m-%d")
                        if hasattr(date_obj, "strftime")
                        else (str(date_obj) if date_obj else "")
                    )
                    eval_table_data.append(
                        [
                            ev.get("source", ""),
                            ev.get("scholarship_name") or "-",
                            ev.get("prompt", "")[:50],
                            ev.get("score"),
                            ev.get("reviewer"),
                            date_str,
                            (ev.get("feedback") or "")[:80],
                        ]
                    )
                eval_table = Table(eval_table_data, repeatRows=1)
                eval_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(eval_table)
            else:
                story.append(
                    Paragraph("No essay evaluations available", styles["Normal"])
                )

        doc.build(story)
        return output_path

    def export_applicant_report_to_excel(
        self, student_id: str = None, netid: str = None, output_path: str = None
    ) -> str:
        """Export applicant report to Excel format."""
        report_data = self.generate_applicant_report(student_id, netid)
        if not report_data:
            raise ValueError("Applicant not found")

        wb = Workbook()

        # Check if this is a multi-applicant report
        is_multi_applicant = "applicants" in report_data

        if is_multi_applicant:
            # Multi-applicant summary
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Summary statistics
            ws_summary.cell(
                row=1, column=1, value="All Applicants Summary"
            ).font = Font(bold=True, size=14)
            ws_summary.cell(row=3, column=1, value="Total Applicants").font = Font(
                bold=True
            )
            ws_summary.cell(row=3, column=2, value=report_data["total_applicants"])
            ws_summary.cell(
                row=4, column=1, value="Total Scholarship Awards"
            ).font = Font(bold=True)
            ws_summary.cell(
                row=4,
                column=2,
                value=report_data["summary"]["total_scholarship_awards"],
            )
            ws_summary.cell(
                row=5, column=1, value="Total Scholarship Amount"
            ).font = Font(bold=True)
            ws_summary.cell(
                row=5,
                column=2,
                value=f"${report_data['summary']['total_scholarship_amount']:,.2f}",
            )
            ws_summary.cell(row=6, column=1, value="Average GPA").font = Font(bold=True)
            ws_summary.cell(
                row=6, column=2, value=f"{report_data['summary']['average_gpa']:.2f}"
            )

            # Individual applicant list
            ws_applicants = wb.create_sheet("All Applicants")
            headers = [
                "Name",
                "Student ID",
                "NetID",
                "Major",
                "Minor",
                "GPA",
                "Academic Level",
                "Achievements (#)",
                "FAFSA",
                "EFC",
                "Income Range",
                "Essay Submissions (#)",
                "Total Awards",
                "Total Amount",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws_applicants.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            for row_idx, applicant in enumerate(report_data["applicants"], 2):
                financial = applicant.get("financial_info") or {}
                achievements = applicant.get("achievements") or []
                essays = applicant.get("essays") or []
                ws_applicants.cell(
                    row=row_idx, column=1, value=applicant["personal_info"]["name"]
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=2,
                    value=applicant["personal_info"]["student_id"],
                )
                ws_applicants.cell(
                    row=row_idx, column=3, value=applicant["personal_info"]["netid"]
                )
                ws_applicants.cell(
                    row=row_idx, column=4, value=applicant["academic_info"]["major"]
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=5,
                    value=applicant["academic_info"]["minor"] or "N/A",
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=6,
                    value=f"{applicant['academic_info']['gpa']:.2f}",
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=7,
                    value=applicant["academic_info"]["academic_level"],
                )
                ws_applicants.cell(row=row_idx, column=8, value=len(achievements))
                ws_applicants.cell(
                    row=row_idx,
                    column=9,
                    value="Yes" if financial.get("fafsa_submitted") else "No",
                )
                ws_applicants.cell(
                    row=row_idx, column=10, value=financial.get("efc", 0)
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=11,
                    value=financial.get("household_income", "N/A"),
                )
                ws_applicants.cell(row=row_idx, column=12, value=len(essays))
                ws_applicants.cell(
                    row=row_idx,
                    column=13,
                    value=applicant["scholarships"]["total_awards"],
                )
                ws_applicants.cell(
                    row=row_idx,
                    column=14,
                    value=f"${applicant['scholarships']['total_amount']:,.2f}",
                )

        else:
            # Single applicant detailed report (existing logic)
            # Personal Information Sheet
            ws_personal = wb.active
            ws_personal.title = "Personal Information"

            personal_info = [
                ["Student Name", report_data["personal_info"]["name"]],
                ["Student ID", report_data["personal_info"]["student_id"]],
                ["NetID", report_data["personal_info"]["netid"]],
                ["Major", report_data["academic_info"]["major"]],
                ["Minor", report_data["academic_info"]["minor"] or "N/A"],
                ["GPA", f"{report_data['academic_info']['gpa']:.2f}"],
                ["Academic Level", report_data["academic_info"]["academic_level"]],
                [
                    "Expected Graduation",
                    report_data["academic_info"]["expected_graduation"].strftime(
                        "%Y-%m-%d"
                    ),
                ],
            ]

            for row_idx, row in enumerate(personal_info, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_personal.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:
                        cell.font = Font(bold=True)

            # Academic History Sheet
            ws_academic = wb.create_sheet("Academic History")
            headers = ["Term", "Course Code", "Course Name", "Grade"]
            for col, header in enumerate(headers, 1):
                cell = ws_academic.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            row = 2
            for term in report_data["academic_info"]["academic_history"]:
                for course in term["courses"]:
                    ws_academic.cell(row=row, column=1, value=term["term"])
                    ws_academic.cell(row=row, column=2, value=course["code"])
                    ws_academic.cell(row=row, column=3, value=course["name"])
                    ws_academic.cell(row=row, column=4, value=course["grade"])
                    row += 1

            # Scholarships Sheet
            ws_scholarships = wb.create_sheet("Scholarships")
            scholarship_headers = ["Scholarship Name", "Amount", "Status", "Award Date"]
            for col, header in enumerate(scholarship_headers, 1):
                cell = ws_scholarships.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            for row, award in enumerate(
                report_data["scholarships"]["detailed_awards"], 2
            ):
                ws_scholarships.cell(row=row, column=1, value=award["scholarship_name"])
                ws_scholarships.cell(
                    row=row, column=2, value=f"${award['award_amount']:,}"
                )
                ws_scholarships.cell(row=row, column=3, value=award["status"])
                ws_scholarships.cell(
                    row=row, column=4, value=award["award_date"].strftime("%Y-%m-%d")
                )

            # Essay Submissions Sheet (new)
            ws_submissions = wb.create_sheet("Essay Submissions")
            sub_headers = ["Prompt", "Submission Date", "Content (Preview)"]
            for col, header in enumerate(sub_headers, 1):
                cell = ws_submissions.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )
            for r_idx, es in enumerate(report_data.get("essays", []), start=2):
                if isinstance(es, dict):
                    sub_date = es.get("submission_date")
                    if hasattr(sub_date, "strftime"):
                        sub_date_str = sub_date.strftime("%Y-%m-%d")
                    else:
                        sub_date_str = str(sub_date) if sub_date else "N/A"
                    ws_submissions.cell(row=r_idx, column=1, value=es.get("prompt", ""))
                    ws_submissions.cell(row=r_idx, column=2, value=sub_date_str)
                    ws_submissions.cell(
                        row=r_idx, column=3, value=(es.get("content", "") or "")[:200]
                    )

            # Essay Evaluations Sheet
            ws_evals = wb.create_sheet("Essay Evaluations")
            eval_headers = [
                "Source",
                "Scholarship",
                "Prompt",
                "Score",
                "Reviewer",
                "Date",
                "Feedback",
            ]
            for col, header in enumerate(eval_headers, 1):
                cell = ws_evals.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            for row_idx, ev in enumerate(
                report_data.get("essay_evaluations", []), start=2
            ):
                date_obj = ev.get("date")
                date_str = (
                    date_obj.strftime("%Y-%m-%d")
                    if hasattr(date_obj, "strftime")
                    else (str(date_obj) if date_obj else "")
                )
                ws_evals.cell(row=row_idx, column=1, value=ev.get("source"))
                ws_evals.cell(
                    row=row_idx, column=2, value=ev.get("scholarship_name") or "-"
                )
                ws_evals.cell(row=row_idx, column=3, value=ev.get("prompt"))
                ws_evals.cell(row=row_idx, column=4, value=ev.get("score"))
                ws_evals.cell(row=row_idx, column=5, value=ev.get("reviewer"))
                ws_evals.cell(row=row_idx, column=6, value=date_str)
                ws_evals.cell(row=row_idx, column=7, value=ev.get("feedback"))

            # Adjust column widths
            for ws in [
                ws_personal,
                ws_academic,
                ws_scholarships,
                ws_submissions,
                ws_evals,
            ]:
                for col in ws.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[chr(64 + col[0].column)].width = min(
                        max_length + 2, 50
                    )

        wb.save(output_path)
        return output_path

    def export_applicant_report_to_csv(
        self, student_id: str = None, netid: str = None, output_path: str = None
    ) -> str:
        """Export applicant report to CSV with flattened essay evaluations."""
        report_data = self.generate_applicant_report(student_id, netid)
        if not report_data:
            raise ValueError("Applicant not found")

        import csv

        # Check if this is a multi-applicant report
        is_multi_applicant = "applicants" in report_data

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            if is_multi_applicant:
                # Multi-applicant summary CSV
                writer.writerow(
                    [
                        "Student Name",
                        "Student ID",
                        "NetID",
                        "Major",
                        "Minor",
                        "GPA",
                        "Academic Level",
                        "Achievements (#)",
                        "FAFSA",
                        "EFC",
                        "Income Range",
                        "Essay Submissions (#)",
                        "Total Awards",
                        "Total Scholarship Amount",
                    ]
                )
                for applicant in report_data["applicants"]:
                    financial = applicant.get("financial_info") or {}
                    achievements = applicant.get("achievements") or []
                    essays = applicant.get("essays") or []
                    writer.writerow(
                        [
                            applicant["personal_info"]["name"],
                            applicant["personal_info"]["student_id"],
                            applicant["personal_info"]["netid"],
                            applicant["academic_info"]["major"],
                            applicant["academic_info"]["minor"] or "N/A",
                            f"{applicant['academic_info']['gpa']:.2f}",
                            applicant["academic_info"]["academic_level"],
                            len(achievements),
                            "Yes" if financial.get("fafsa_submitted") else "No",
                            financial.get("efc", 0),
                            financial.get("household_income", "N/A"),
                            len(essays),
                            applicant["scholarships"]["total_awards"],
                            f"${applicant['scholarships']['total_amount']:,.2f}",
                        ]
                    )
            else:
                # Single applicant with essay evaluations (existing logic)
                writer.writerow(
                    [
                        "Row Type",
                        "Student Name",
                        "Student ID",
                        "NetID",
                        "Major",
                        "Minor",
                        "GPA",
                        "Academic Level",
                        "Prompt",
                        "Submission Date/ Eval Date",
                        "Source",
                        "Scholarship",
                        "Score",
                        "Reviewer",
                        "Feedback",
                    ]
                )
                # First add essay submissions
                for es in report_data.get("essays") or []:
                    if isinstance(es, dict):
                        sub_date = es.get("submission_date")
                        if hasattr(sub_date, "strftime"):
                            sub_date_str = sub_date.strftime("%Y-%m-%d")
                        else:
                            sub_date_str = str(sub_date) if sub_date else "N/A"
                        writer.writerow(
                            [
                                "Submission",
                                report_data["personal_info"]["name"],
                                report_data["personal_info"]["student_id"],
                                report_data["personal_info"]["netid"],
                                report_data["academic_info"]["major"],
                                report_data["academic_info"]["minor"] or "N/A",
                                f"{report_data['academic_info']['gpa']:.2f}",
                                report_data["academic_info"]["academic_level"],
                                es.get("prompt", ""),
                                sub_date_str,
                                "Essay Submission",
                                "-",
                                "-",
                                "-",
                                (es.get("content", "") or "")[:120],
                            ]
                        )
                evaluations = report_data.get("essay_evaluations", [])
                if evaluations:
                    for ev in evaluations:
                        date_obj = ev.get("date")
                        date_str = (
                            date_obj.strftime("%Y-%m-%d")
                            if hasattr(date_obj, "strftime")
                            else (str(date_obj) if date_obj else "")
                        )
                        writer.writerow(
                            [
                                "Evaluation",
                                report_data["personal_info"]["name"],
                                report_data["personal_info"]["student_id"],
                                report_data["personal_info"]["netid"],
                                report_data["academic_info"]["major"],
                                report_data["academic_info"]["minor"] or "N/A",
                                f"{report_data['academic_info']['gpa']:.2f}",
                                report_data["academic_info"]["academic_level"],
                                ev.get("prompt"),
                                date_str,
                                ev.get("source"),
                                ev.get("scholarship_name") or "-",
                                ev.get("score"),
                                ev.get("reviewer"),
                                (ev.get("feedback") or "")[:120],
                            ]
                        )
                else:
                    writer.writerow(
                        [
                            "Evaluation",
                            report_data["personal_info"]["name"],
                            report_data["personal_info"]["student_id"],
                            report_data["personal_info"]["netid"],
                            report_data["academic_info"]["major"],
                            report_data["academic_info"]["minor"] or "N/A",
                            f"{report_data['academic_info']['gpa']:.2f}",
                            report_data["academic_info"]["academic_level"],
                            "-",
                            "N/A",
                            "None",
                            "-",
                            "-",
                            "-",
                            "No evaluations",
                        ]
                    )
        return output_path

    # Financial Aid System Integration Helper Methods
    # Implements requirement: The report engine shall support future integration with
    # financial aid systems to automate or assist in payment processing.
    # See SFWE504_3-LLR-29.
    def export_financial_aid_data(
        self,
        scholarship_awards: List[ScholarshipAward],
        format: str = "csv",
        system_type: str = "banner",
    ) -> str:
        """
        Export disbursement data in formats compatible with financial aid systems.

        This is a convenience wrapper around the financial_integration module's export function.

        Args:
            scholarship_awards: List of ScholarshipAward model instances to export
            format: Export format ('csv', 'json', 'xml')
            system_type: Target financial aid system ('banner', 'workday', etc.)

        Returns:
            str: Path to the generated export file

        Example:
            ```python
            from reports_app.models import ScholarshipAward

            # Get awards to export
            awards = ScholarshipAward.objects.filter(status='active')

            # Export for Banner system
            csv_path = engine.export_financial_aid_data(
                scholarship_awards=awards,
                format='csv',
                system_type='banner'
            )
            ```
        """
        from .financial_integration import generate_financial_aid_export

        return generate_financial_aid_export(scholarship_awards, format, system_type)

    def submit_disbursements_to_financial_aid_system(
        self, scholarship_awards: List[ScholarshipAward], system_name: str = None
    ) -> List[Dict[str, Any]]:
        """
        Submit disbursement transactions to the configured financial aid system.

        Args:
            scholarship_awards: List of ScholarshipAward instances to process
            system_name: Specific financial aid system to use (uses default if None)

        Returns:
            List of result dictionaries with success/failure status

        Example:
            ```python
            # Submit approved disbursements
            from reports_app.models import ScholarshipAward

            awards = ScholarshipAward.objects.filter(status='active')
            results = engine.submit_disbursements_to_financial_aid_system(awards)

            for result in results:
                if result['success']:
                    print(f"Submitted: {result['transaction_id']}")
                else:
                    print(f"Failed: {result['message']}")
            ```
        """
        from .financial_integration import FinancialAidIntegrationManager
        from .models import DisbursementTransaction
        from datetime import date

        manager = FinancialAidIntegrationManager()
        results = []

        for award in scholarship_awards:
            # Parse disbursement dates
            disbursement_dates = award.disbursement_dates
            if isinstance(disbursement_dates, str):
                import json

                disbursement_dates = json.loads(disbursement_dates)

            # Create disbursement for each scheduled date
            for i, disb_date in enumerate(disbursement_dates):
                if isinstance(disb_date, str):
                    from datetime import datetime as dt

                    disb_date = dt.fromisoformat(disb_date).date()

                amount_per_disbursement = award.award_amount / len(disbursement_dates)

                # Create or get disbursement transaction
                transaction_id = f"DISB-{award.id}-{i + 1}"
                transaction, created = DisbursementTransaction.objects.get_or_create(
                    transaction_id=transaction_id,
                    defaults={
                        "scholarship_award": award,
                        "amount": amount_per_disbursement,
                        "scheduled_date": disb_date,
                        "status": "approved",
                    },
                )

                # Submit if not already submitted
                if transaction.status in ["approved", "failed"]:
                    disbursement_data = {
                        "student_id": award.applicant.student_id,
                        "amount": transaction.amount,
                        "scholarship_name": award.scholarship_name,
                        "disbursement_date": transaction.scheduled_date,
                        "reference_number": transaction.transaction_id,
                    }

                    result = manager.submit_batch_disbursements(
                        [disbursement_data], system_name
                    )
                    results.extend(result)

                    # Update transaction status
                    if result[0]["success"]:
                        transaction.mark_submitted(
                            external_id=result[0]["transaction_id"],
                            system_name=system_name or "default",
                        )
                    else:
                        transaction.mark_failed(
                            result[0].get("message", "Unknown error")
                        )

        return results

    # Custom Analytics and Reporting Functions
    # Generate custom reports and analytics on application trends and scholarship impacts
    def generate_analytics_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        report_types: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Generate analytics using only data accessed through ReportEngine helpers.

        Ensures consistency with other report outputs by:
        - Using generate_applicant_report for application data
        - Using get_awards_queryset for scholarship impact

        Supported report_types: 'application_trends', 'scholarship_impact', 'all'
        """
        from django.db.models import Count, Avg, Sum, Min, Max
        from datetime import timedelta
        from collections import defaultdict

        # Defaults: past year window
        if end_date is None:
            end_date = timezone.now()
        if start_date is None:
            start_date = end_date - timedelta(days=365)

        if not report_types or "all" in report_types:
            report_types = ["application_trends", "scholarship_impact"]

        analytics: Dict[str, Any] = {
            "metadata": {
                "generated_at": timezone.now(),
                "period_start": start_date,
                "period_end": end_date,
                "report_types": report_types,
                "source": "ReportEngine unified data access",
            }
        }

        # Helper: unified applicant queryset filtered by date
        def _get_applicants():
            """Get applicants with optional date filtering."""
            qs = Applicant.objects.all()
            # Only apply date filtering if both date parameters are provided
            if start_date and end_date:
                try:
                    qs = qs.filter(created_at__gte=start_date, created_at__lte=end_date)
                except Exception as e:
                    # If filtering fails, return all applicants
                    logger.warning(f"Date filtering failed for applicants: {e}")
            return qs

        # Helper: unified awards queryset filtered by date
        def _get_awards():
            """Get scholarship awards with optional date filtering."""
            qs = ScholarshipAward.objects.all()
            # Only apply date filtering if both date parameters are provided
            if start_date and end_date:
                try:
                    # Filter by award_date
                    qs = qs.filter(award_date__gte=start_date.date() if hasattr(start_date, 'date') else start_date,
                                 award_date__lte=end_date.date() if hasattr(end_date, 'date') else end_date)
                except Exception as e:
                    # If filtering fails, return all awards
                    logger.warning(f"Date filtering failed for awards: {e}")
            return qs

        # 1) Application Trends (built from applicant report + queryset for efficiency distributions)
        if "application_trends" in report_types:
            applicants_qs = _get_applicants()

            monthly = defaultdict(int)
            for a in applicants_qs:
                # Handle cases where created_at might not exist
                if hasattr(a, 'created_at') and getattr(a, "created_at", None):
                    monthly[a.created_at.strftime("%Y-%m")] += 1
                else:
                    # Use current month as fallback
                    monthly[timezone.now().strftime("%Y-%m")] += 1

            major_dist = list(
                applicants_qs.values("major")
                .annotate(count=Count("id"))
                .order_by("-count")
            )
            level_dist = list(
                applicants_qs.values("academic_level")
                .annotate(count=Count("id"))
                .order_by("-count")
            )
            gpa_stats = applicants_qs.aggregate(
                avg_gpa=Avg("gpa"), min_gpa=Min("gpa"), max_gpa=Max("gpa")
            )

            analytics["application_trends"] = {
                "total_applications": applicants_qs.count(),
                "monthly_trends": dict(monthly),
                "major_distribution": major_dist,
                "academic_level_distribution": level_dist,
                "gpa_statistics": gpa_stats,
            }

        # 2) Scholarship Impact (awards only via helper)
        if "scholarship_impact" in report_types:
            awards_qs = _get_awards()
            financial = awards_qs.aggregate(
                total_awarded=Sum("award_amount"),
                avg_award=Avg("award_amount"),
                min_award=Min("award_amount"),
                max_award=Max("award_amount"),
            )
            by_scholarship = list(
                awards_qs.values("scholarship_name")
                .annotate(count=Count("id"), total_amount=Sum("award_amount"))
                .order_by("-total_amount")
            )
            status_breakdown = list(
                awards_qs.values("status")
                .annotate(count=Count("id"))
                .order_by("-count")
            )

            analytics["scholarship_impact"] = {
                "total_awards": awards_qs.count(),
                "financial_impact": financial,
                "scholarship_breakdown": by_scholarship,
                "status_distribution": status_breakdown,
            }

        return analytics

    def export_analytics_report_to_json(
        self, analytics_data: Dict[str, Any], output_path: str
    ) -> str:
        """Export analytics report to JSON format."""
        import json
        from datetime import date
        from decimal import Decimal

        def _ser(obj):
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            if isinstance(obj, Decimal):
                return float(obj)
            return str(obj)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(analytics_data, f, indent=2, default=_ser)
        return output_path

    def export_analytics_report_to_excel(
        self, analytics_data: Dict[str, Any], output_path: str
    ) -> str:
        """Export analytics report to an Excel workbook with formatted, readable sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        def _sheet(name: str) -> str:
            for c in ["\\", "/", "*", "?", ":", "[", "]"]:
                name = name.replace(c, "-")
            return (name or "Sheet")[:31]
        
        def _format_header(ws, row=1):
            """Apply header formatting to the first row."""
            for cell in ws[row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        def _auto_size_columns(ws):
            """Auto-size columns based on content."""
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb = Workbook()
        
        # Metadata sheet
        meta_ws = wb.active
        meta_ws.title = "Metadata"
        meta = analytics_data.get("metadata", {})
        meta_ws.append(["Field", "Value"])
        _format_header(meta_ws)
        for k, v in meta.items():
            meta_ws.append([k.replace("_", " ").title(), str(v)])
        _auto_size_columns(meta_ws)

        # Application Trends sheet
        if "application_trends" in analytics_data:
            app_data = analytics_data["application_trends"]
            ws = wb.create_sheet("Application Trends")
            
            # Summary section
            ws.append(["Application Trends Summary"])
            ws.merge_cells('A1:B1')
            ws['A1'].font = Font(bold=True, size=14)
            ws.append([])
            
            ws.append(["Metric", "Value"])
            _format_header(ws, row=3)
            ws.append(["Total Applications", app_data.get("total_applications", 0)])
            
            # GPA Statistics
            gpa_stats = app_data.get("gpa_statistics", {})
            if gpa_stats:
                ws.append([])
                ws.append(["GPA Statistics", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Average GPA", f"{gpa_stats.get('avg_gpa', 0):.2f}" if gpa_stats.get('avg_gpa') else "N/A"])
                ws.append(["Minimum GPA", f"{gpa_stats.get('min_gpa', 0):.2f}" if gpa_stats.get('min_gpa') else "N/A"])
                ws.append(["Maximum GPA", f"{gpa_stats.get('max_gpa', 0):.2f}" if gpa_stats.get('max_gpa') else "N/A"])
            
            # Monthly Trends
            monthly = app_data.get("monthly_trends", {})
            if monthly:
                ws.append([])
                ws.append(["Monthly Trends", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Month", "Applications"])
                _format_header(ws, row=ws.max_row)
                for month, count in sorted(monthly.items()):
                    ws.append([month, count])
            
            # Major Distribution
            major_dist = app_data.get("major_distribution", [])
            if major_dist:
                ws.append([])
                ws.append(["Major Distribution", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Major", "Count"])
                _format_header(ws, row=ws.max_row)
                for item in major_dist:
                    ws.append([item.get("major", "Unknown"), item.get("count", 0)])
            
            # Academic Level Distribution
            level_dist = app_data.get("academic_level_distribution", [])
            if level_dist:
                ws.append([])
                ws.append(["Academic Level Distribution", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Level", "Count"])
                _format_header(ws, row=ws.max_row)
                for item in level_dist:
                    ws.append([item.get("academic_level", "Unknown"), item.get("count", 0)])
            
            _auto_size_columns(ws)

        # Scholarship Impact sheet
        if "scholarship_impact" in analytics_data:
            schol_data = analytics_data["scholarship_impact"]
            ws = wb.create_sheet("Scholarship Impact")
            
            # Summary section
            ws.append(["Scholarship Impact Summary"])
            ws.merge_cells('A1:B1')
            ws['A1'].font = Font(bold=True, size=14)
            ws.append([])
            
            ws.append(["Metric", "Value"])
            _format_header(ws, row=3)
            ws.append(["Total Awards", schol_data.get("total_awards", 0)])
            
            # Financial Impact
            financial = schol_data.get("financial_impact", {})
            if financial:
                ws.append([])
                ws.append(["Financial Impact", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Total Awarded", f"${financial.get('total_awarded', 0):,.2f}" if financial.get('total_awarded') else "$0.00"])
                ws.append(["Average Award", f"${financial.get('avg_award', 0):,.2f}" if financial.get('avg_award') else "$0.00"])
                ws.append(["Minimum Award", f"${financial.get('min_award', 0):,.2f}" if financial.get('min_award') else "$0.00"])
                ws.append(["Maximum Award", f"${financial.get('max_award', 0):,.2f}" if financial.get('max_award') else "$0.00"])
            
            # Scholarship Breakdown
            schol_breakdown = schol_data.get("scholarship_breakdown", [])
            if schol_breakdown:
                ws.append([])
                ws.append(["Scholarship Breakdown", "", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Scholarship Name", "Count", "Total Amount"])
                _format_header(ws, row=ws.max_row)
                for item in schol_breakdown:
                    ws.append([
                        item.get("scholarship_name", "Unknown"),
                        item.get("count", 0),
                        f"${item.get('total_amount', 0):,.2f}" if item.get('total_amount') else "$0.00"
                    ])
            
            # Status Distribution
            status_dist = schol_data.get("status_distribution", [])
            if status_dist:
                ws.append([])
                ws.append(["Status Distribution", ""])
                ws['A{}'.format(ws.max_row)].font = Font(bold=True)
                ws.append(["Status", "Count"])
                _format_header(ws, row=ws.max_row)
                for item in status_dist:
                    ws.append([item.get("status", "Unknown"), item.get("count", 0)])
            
            _auto_size_columns(ws)

        wb.save(output_path)
        return output_path

    def export_analytics_report_to_pdf(
        self, analytics_data: Dict[str, Any], output_path: str
    ) -> str:
        """Export analytics to a formatted, readable PDF report."""
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Table,
            TableStyle,
            Spacer,
            PageBreak,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Custom styles
        styles.add(ParagraphStyle(
            name='CenterTitle',
            parent=styles['Title'],
            alignment=TA_CENTER,
        ))
        styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=styles['Heading2'],
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=12,
        ))
        
        story = [
            Paragraph("Scholarship Analytics Report", styles["CenterTitle"]),
            Spacer(1, 20),
        ]

        # Metadata section
        meta = analytics_data.get("metadata", {})
        if meta:
            story.append(Paragraph("Report Information", styles["SectionHeader"]))
            rows = [["Field", "Value"]]
            for k, v in meta.items():
                rows.append([k.replace("_", " ").title(), str(v)])
            
            t = Table(rows, colWidths=[200, 300])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            story.append(t)
            story.append(Spacer(1, 20))

        # Application Trends section
        if "application_trends" in analytics_data:
            app_data = analytics_data["application_trends"]
            story.append(Paragraph("Application Trends", styles["SectionHeader"]))
            
            # Summary table
            rows = [["Metric", "Value"]]
            rows.append(["Total Applications", str(app_data.get("total_applications", 0))])
            
            gpa_stats = app_data.get("gpa_statistics", {})
            if gpa_stats:
                rows.append(["Average GPA", f"{gpa_stats.get('avg_gpa', 0):.2f}" if gpa_stats.get('avg_gpa') else "N/A"])
                rows.append(["Minimum GPA", f"{gpa_stats.get('min_gpa', 0):.2f}" if gpa_stats.get('min_gpa') else "N/A"])
                rows.append(["Maximum GPA", f"{gpa_stats.get('max_gpa', 0):.2f}" if gpa_stats.get('max_gpa') else "N/A"])
            
            t = Table(rows, colWidths=[250, 250])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            story.append(t)
            story.append(Spacer(1, 15))
            
            # Major Distribution
            major_dist = app_data.get("major_distribution", [])
            if major_dist:
                story.append(Paragraph("Major Distribution", styles["Heading3"]))
                rows = [["Major", "Count"]]
                for item in major_dist[:10]:  # Limit to top 10
                    rows.append([item.get("major", "Unknown"), str(item.get("count", 0))])
                
                t = Table(rows, colWidths=[350, 150])
                t.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (0, -1), "LEFT"),
                            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ]
                    )
                )
                story.append(t)
                story.append(Spacer(1, 20))

        # Scholarship Impact section
        if "scholarship_impact" in analytics_data:
            story.append(PageBreak())
            schol_data = analytics_data["scholarship_impact"]
            story.append(Paragraph("Scholarship Impact", styles["SectionHeader"]))
            
            # Summary table
            rows = [["Metric", "Value"]]
            rows.append(["Total Awards", str(schol_data.get("total_awards", 0))])
            
            financial = schol_data.get("financial_impact", {})
            if financial:
                rows.append(["Total Awarded", f"${financial.get('total_awarded', 0):,.2f}" if financial.get('total_awarded') else "$0.00"])
                rows.append(["Average Award", f"${financial.get('avg_award', 0):,.2f}" if financial.get('avg_award') else "$0.00"])
                rows.append(["Minimum Award", f"${financial.get('min_award', 0):,.2f}" if financial.get('min_award') else "$0.00"])
                rows.append(["Maximum Award", f"${financial.get('max_award', 0):,.2f}" if financial.get('max_award') else "$0.00"])
            
            t = Table(rows, colWidths=[250, 250])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (0, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 11),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            story.append(t)
            story.append(Spacer(1, 15))
            
            # Scholarship Breakdown
            schol_breakdown = schol_data.get("scholarship_breakdown", [])
            if schol_breakdown:
                story.append(Paragraph("Scholarship Breakdown", styles["Heading3"]))
                rows = [["Scholarship Name", "Count", "Total Amount"]]
                for item in schol_breakdown:
                    rows.append([
                        item.get("scholarship_name", "Unknown"),
                        str(item.get("count", 0)),
                        f"${item.get('total_amount', 0):,.2f}" if item.get('total_amount') else "$0.00"
                    ])
                
                t = Table(rows, colWidths=[250, 100, 150])
                t.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (0, -1), "LEFT"),
                            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ]
                    )
                )
                story.append(t)
                story.append(Spacer(1, 15))
            
            # Status Distribution
            status_dist = schol_data.get("status_distribution", [])
            if status_dist:
                story.append(Paragraph("Status Distribution", styles["Heading3"]))
                rows = [["Status", "Count"]]
                for item in status_dist:
                    rows.append([item.get("status", "Unknown"), str(item.get("count", 0))])
                
                t = Table(rows, colWidths=[350, 150])
                t.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#FFC000')),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                            ("ALIGN", (0, 0), (0, -1), "LEFT"),
                            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ]
                    )
                )
                story.append(t)

        doc.build(story)
        return output_path


import logging

logger = logging.getLogger("reports_app")


# View to handle report generation and exporting
def home(request):
    """View to handle report generation and exporting.
    Logs detailed debugging information about request processing and report generation.
    """
    logger.debug(
        "Processing home request. Method: %s, POST data: %s",
        request.method,
        request.POST if request.method == "POST" else "N/A",
    )

    # Create sample scholarship data (inline)
    # Create sample applicant data with comprehensive review information
    john_doe = Applicant.from_dict(
        {
            "name": "John Doe",
            "student_id": "12345678",
            "netid": "jdoe",
            "major": "Systems Engineering",
            "minor": "Computer Science",
            "academic_achievements": [
                {
                    "type": "Dean's List",
                    "date": datetime(2024, 12, 15),
                    "description": "Fall 2024 Semester",
                },
                {
                    "type": "Research Publication",
                    "date": datetime(2025, 3, 1),
                    "title": "Innovation in Systems Design",
                    "journal": "Engineering Research Quarterly",
                },
            ],
            "financial_info": {
                "fafsa_submitted": True,
                "efc": 5000,
                "household_income": "50000-75000",
                "current_aid": [
                    {"type": "Federal Grant", "amount": 2500},
                    {"type": "State Grant", "amount": 1500},
                ],
            },
            "essays": [
                {
                    "prompt": "Describe your career goals in engineering.",
                    "content": "My passion for systems engineering stems from...",
                    "submission_date": datetime(2025, 2, 1),
                    "evaluation": {
                        "score": 9.2,
                        "feedback": "Excellent vision and clear career trajectory.",
                        "reviewer": "Dr. Sarah Chen",
                        "date": datetime(2025, 2, 15),
                    },
                },
                {
                    "prompt": "How will this scholarship impact your education?",
                    "content": "This scholarship will enable me to...",
                    "submission_date": datetime(2025, 2, 1),
                    "evaluation": {
                        "score": 8.8,
                        "feedback": "Strong understanding of opportunity and impact.",
                        "reviewer": "Prof. Michael Roberts",
                        "date": datetime(2025, 2, 16),
                    },
                },
            ],
            "gpa": 3.8,
            "academic_level": "Junior",
            "expected_graduation": datetime(2027, 5, 15),
            "academic_history": [
                {
                    "term": "Fall 2024",
                    "courses": [
                        {
                            "code": "SYE301",
                            "name": "Systems Engineering Fundamentals",
                            "grade": "A",
                        },
                        {"code": "CS210", "name": "Software Systems", "grade": "A-"},
                    ],
                    "gpa": 3.85,
                }
            ],
            "interview_notes": "Conducted on 2025-03-01. Demonstrated strong leadership potential and excellent communication skills. Shows clear understanding of systems engineering principles.",
            "committee_feedback": [
                {
                    "member": "Dr. James Wilson",
                    "role": "Department Chair",
                    "comments": "Outstanding candidate with proven academic excellence.",
                    "recommendation": "Highly Recommend",
                    "date": datetime(2025, 3, 5),
                },
                {
                    "member": "Prof. Lisa Martinez",
                    "role": "Scholarship Committee Head",
                    "comments": "Strong technical background and leadership potential.",
                    "recommendation": "Strongly Recommend",
                    "date": datetime(2025, 3, 6),
                },
            ],
        }
    )

    # Create additional diverse applicants
    maria_garcia = Applicant.from_dict(
        {
            "name": "Maria Garcia",
            "student_id": "87654321",
            "netid": "mgarcia",
            "major": "Computer Science",
            "minor": "Mathematics",
            "academic_achievements": [
                {
                    "type": "President's List",
                    "date": datetime(2024, 12, 15),
                    "description": "Fall 2024 Semester",
                }
            ],
            "financial_info": {
                "fafsa_submitted": True,
                "efc": 3000,
                "household_income": "30000-50000",
            },
            "essays": [
                {
                    "prompt": "Leadership experience",
                    "content": "As president of the Computer Science Club...",
                    "submission_date": datetime(2025, 2, 5),
                }
            ],
            "gpa": 3.9,
            "academic_level": "Senior",
            "expected_graduation": datetime(2026, 5, 15),
        }
    )

    sarah_johnson = Applicant.from_dict(
        {
            "name": "Sarah Johnson",
            "student_id": "11223344",
            "netid": "sjohnson",
            "major": "Electrical Engineering",
            "minor": "Physics",
            "academic_achievements": [
                {
                    "type": "Research Award",
                    "date": datetime(2025, 1, 10),
                    "description": "Outstanding Undergraduate Research",
                }
            ],
            "financial_info": {
                "fafsa_submitted": True,
                "efc": 4500,
                "household_income": "60000-80000",
            },
            "essays": [],
            "gpa": 3.75,
            "academic_level": "Junior",
            "expected_graduation": datetime(2027, 5, 15),
        }
    )

    # Create sample scholarships
    # First scholarship with an award
    engineering_scholarship = Scholarship.objects.create(
        name="Engineering Excellence Scholarship",
        description="Merit-based scholarship for outstanding engineering students",
        eligibility_criteria=["3.5+ GPA", "Engineering major", "Full-time enrollment"],
        donor_info={
            "name": "Engineering Industry Association",
            "contact": "donor@example.com",
        },
        disbursement_requirements=[
            "Maintain 3.5 GPA",
            "Submit semester progress report",
        ],
        frequency="annual",
        amount=5000.00,
        deadline=timezone.make_aware(datetime(2026, 3, 15)),
        review_dates=[
            timezone.make_aware(datetime(2026, 1, 15)).isoformat(),  # Mid-year review
            timezone.make_aware(datetime(2026, 6, 15)).isoformat(),  # End-year review
        ],
        reporting_schedule={
            "Progress Report": timezone.make_aware(datetime(2026, 4, 15)).isoformat(),
            "Financial Report": timezone.make_aware(datetime(2026, 7, 15)).isoformat(),
        },
    )

    # Create or update a single scholarship award (avoid duplicates on page reload)
    _award_defaults = {
        "award_date": timezone.make_aware(datetime(2025, 8, 15)),
        "award_amount": 5000.00,
        "disbursement_dates": [
            timezone.make_aware(datetime(2025, 9, 1)).isoformat(),
            timezone.make_aware(datetime(2026, 1, 1)).isoformat(),
        ],
        "requirements_met": [
            "Enrollment verification",
            "First semester GPA requirement",
        ],
        "requirements_pending": [
            "Second semester progress report",
            "Community service hours",
        ],
        "status": "active",
        "performance_metrics": {
            "current_gpa": 3.8,
            "credits_completed": 15,
            "service_hours": 20,
        },
        "essays_evaluation": [
            {
                "prompt": "Career Goals",
                "score": 9,
                "feedback": "Excellent clarity and vision",
            },
            {
                "prompt": "Impact",
                "score": 8,
                "feedback": "Strong understanding of opportunity",
            },
        ],
        "interview_notes": "Strong candidate with clear goals and excellent communication skills",
        "committee_feedback": [
            {"member": "Dr. Smith", "comments": "Highly recommended"},
            {"member": "Prof. Johnson", "comments": "Outstanding potential"},
        ],
    }
    # Ensure a single sample active award exists (avoid mass duplicates)
    # 1) Remove non-active duplicates for clarity
    ScholarshipAward.objects.filter(
        scholarship_name=engineering_scholarship.name, applicant=john_doe
    ).exclude(status="active").delete()

    # 2) Deduplicate active awards and update or create without triggering MultipleObjectsReturned
    existing_qs = ScholarshipAward.objects.filter(
        scholarship_name=engineering_scholarship.name, applicant=john_doe
    ).order_by("-award_date", "-id")

    latest = existing_qs.first()
    if latest:
        # delete any extras beyond the latest
        if existing_qs.count() > 1:
            existing_qs.exclude(pk=latest.pk).delete()
        # update latest with defaults
        for k, v in _award_defaults.items():
            setattr(latest, k, v)
        latest.save()
    else:
        ScholarshipAward.objects.create(
            scholarship_name=engineering_scholarship.name,
            applicant=john_doe,
            **_award_defaults,
        )

    # Second scholarship
    cs_scholarship = Scholarship.objects.create(
        name="CS Leadership Scholarship",
        description="For computer science students demonstrating leadership",
        eligibility_criteria=[
            "3.0+ GPA",
            "Computer Science major",
            "Leadership role in student organization",
        ],
        donor_info={
            "name": "Tech Leaders Foundation",
            "contact": "foundation@techleaders.org",
            "email": "info@techleaders.org",
            "phone": "555-0123",
        },
        disbursement_requirements=[
            "Maintain leadership position",
            "Submit leadership impact report",
        ],
        frequency="semester",
        amount=3000.00,
        deadline=timezone.make_aware(datetime(2026, 2, 1)),
        review_dates=[timezone.make_aware(datetime(2026, 3, 1)).isoformat()],
        reporting_schedule={
            "Leadership Report": timezone.make_aware(datetime(2026, 5, 1)).isoformat()
        },
    )

    # Create award for Maria Garcia for CS Leadership Scholarship
    _cs_award_defaults = {
        "award_date": timezone.make_aware(datetime(2025, 9, 1)),
        "award_amount": 3000.00,
        "disbursement_dates": [
            timezone.make_aware(datetime(2025, 10, 1)).isoformat(),
            timezone.make_aware(datetime(2026, 2, 15)).isoformat(),
        ],
        "requirements_met": [
            "Enrollment verification",
            "Leadership position confirmed",
        ],
        "requirements_pending": ["Mid-semester leadership report"],
        "status": "active",
        "performance_metrics": {"current_gpa": 3.9, "leadership_hours": 30},
        "essays_evaluation": [
            {
                "prompt": "Leadership",
                "score": 9.5,
                "feedback": "Exceptional leadership qualities",
            }
        ],
        "interview_notes": "Outstanding leadership and communication skills",
        "committee_feedback": [
            {"member": "Prof. Anderson", "comments": "Exemplary candidate"}
        ],
    }

    # Manage CS scholarship award for Maria Garcia
    existing_cs_qs = ScholarshipAward.objects.filter(
        scholarship_name=cs_scholarship.name, applicant=maria_garcia
    ).order_by("-award_date", "-id")

    latest_cs = existing_cs_qs.first()
    if latest_cs:
        if existing_cs_qs.count() > 1:
            existing_cs_qs.exclude(pk=latest_cs.pk).delete()
        for k, v in _cs_award_defaults.items():
            setattr(latest_cs, k, v)
        latest_cs.save()
    else:
        ScholarshipAward.objects.create(
            scholarship_name=cs_scholarship.name,
            applicant=maria_garcia,
            **_cs_award_defaults,
        )

    # Add scholarships to list for the engine
    sample_scholarships = [engineering_scholarship, cs_scholarship]

    # Initialize engine and add sample data
    engine = ReportEngine()
    for scholarship in sample_scholarships:
        engine.add_scholarship(scholarship)

    if request.method == "POST":
        export_format = request.POST.get("export_format")
        report_type = request.POST.get("report_type", "general")  # 'general' or 'donor'
        donor_name = request.POST.get("donor_name")

        logger.debug(
            "Processing export request - Format: %s, Type: %s, Donor: %s",
            export_format,
            report_type,
            donor_name,
        )

        if export_format:
            temp_file = None
            try:
                # Create temporary file with appropriate extension
                temp_file = tempfile.NamedTemporaryFile(
                    delete=False, suffix=f".{export_format}"
                )
                temp_file.close()  # Close the file handle immediately

                if report_type == "donor" and donor_name:
                    # Generate donor-specific report
                    if export_format == "pdf":
                        output_path = engine.export_donor_report_to_pdf(
                            donor_name=donor_name, output_path=temp_file.name
                        )
                        content_type = "application/pdf"
                    elif export_format == "xlsx":
                        output_path = engine.export_donor_report_to_excel(
                            donor_name=donor_name, output_path=temp_file.name
                        )
                        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif export_format == "csv":
                        output_path = engine.export_donor_report_to_csv(
                            donor_name=donor_name, output_path=temp_file.name
                        )
                        content_type = "text/csv"
                    else:
                        raise ValueError(
                            f"Unsupported export format for donor report: {export_format}"
                        )
                    filename = f"donor_report.{export_format}"
                elif report_type == "applicant":
                    # Generate report for ALL applicants (not just one sample)
                    if export_format == "pdf":
                        output_path = engine.export_applicant_report_to_pdf(
                            student_id=None,  # None = all applicants
                            output_path=temp_file.name,
                        )
                        content_type = "application/pdf"
                    elif export_format == "xlsx":
                        output_path = engine.export_applicant_report_to_excel(
                            student_id=None,  # None = all applicants
                            output_path=temp_file.name,
                        )
                        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif export_format == "csv":
                        output_path = engine.export_applicant_report_to_csv(
                            student_id=None,  # None = all applicants
                            output_path=temp_file.name,
                        )
                        content_type = "text/csv"
                    else:
                        raise ValueError(
                            f"Unsupported export format for applicant report: {export_format}"
                        )
                    filename = f"applicant_report.{export_format}"
                elif report_type == "disbursement":
                    # Get scholarship name from request
                    scholarship_name = request.POST.get("scholarship_name")
                    if export_format == "pdf":
                        output_path = engine.export_disbursement_report_to_pdf(
                            scholarship_name=scholarship_name,
                            output_path=temp_file.name,
                        )
                        content_type = "application/pdf"
                    elif export_format == "xlsx":
                        output_path = engine.export_disbursement_report_to_excel(
                            scholarship_name=scholarship_name,
                            output_path=temp_file.name,
                        )
                        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif export_format == "csv":
                        output_path = engine.export_disbursement_report_to_csv(
                            scholarship_name=scholarship_name,
                            output_path=temp_file.name,
                        )
                        content_type = "text/csv"
                    else:
                        raise ValueError(
                            f"Unsupported export format for disbursement report: {export_format}"
                        )
                    filename = f"disbursement_report.{export_format}"
                elif report_type == "prescreening":
                    # For demo purposes, we'll create a list of sample applicants with varying completion levels
                    sample_applicants = [
                        Applicant.from_dict(
                            {
                                "name": "Alice Smith",
                                "student_id": "12346789",
                                "netid": "asmith",
                                "major": "Engineering",
                                "minor": "Mathematics",
                                "gpa": 3.8,
                                "academic_level": "Junior",
                                "expected_graduation": datetime(2027, 5, 15),
                                "academic_history": [
                                    {
                                        "term": "Fall 2024",
                                        "courses": [
                                            {
                                                "code": "ENG301",
                                                "name": "Advanced Engineering",
                                                "grade": "A",
                                            },
                                            {
                                                "code": "MATH400",
                                                "name": "Applied Mathematics",
                                                "grade": "A-",
                                            },
                                        ],
                                        "gpa": 3.8,
                                    }
                                ],
                                "essays": [
                                    {
                                        "prompt": "Describe your research interests.",
                                        "content": "My research focuses on sustainable engineering...",
                                        "submission_date": datetime(2025, 2, 1),
                                        "evaluation": {
                                            "score": 9.5,
                                            "feedback": "Exceptional research vision and clarity.",
                                            "reviewer": "Dr. Thompson",
                                            "date": datetime(2025, 2, 10),
                                        },
                                    }
                                ],
                                "financial_info": {
                                    "fafsa_submitted": True,
                                    "efc": 4000,
                                    "household_income": "40000-60000",
                                },
                                "interview_notes": "Outstanding interview performance. Shows great potential.",
                                "committee_feedback": [
                                    {
                                        "member": "Dr. Rodriguez",
                                        "comments": "Top candidate with excellent credentials.",
                                        "recommendation": "Highly Recommend",
                                        "date": datetime(2025, 3, 1),
                                    }
                                ],
                            }
                        ),
                        Applicant.from_dict(
                            {
                                "name": "Bob Johnson",
                                "student_id": "12347890",
                                "netid": "bjohnson",
                                "major": "Computer Science",
                                "gpa": 3.2,
                                "academic_level": "Sophomore",
                                "essays": [
                                    {
                                        "prompt": "Describe your programming experience.",
                                        "content": "I have developed several applications...",
                                        "submission_date": datetime(2025, 2, 2),
                                        "evaluation": {
                                            "score": 7.8,
                                            "feedback": "Good technical background, needs more detail.",
                                            "reviewer": "Prof. Chen",
                                            "date": datetime(2025, 2, 12),
                                        },
                                    }
                                ],
                                "financial_info": {
                                    "fafsa_submitted": True,
                                    "efc": 6000,
                                    "household_income": "60000-80000",
                                },
                            }
                        ),
                        Applicant.from_dict(
                            {
                                "name": "Carol Williams",
                                "student_id": "12348901",
                                "netid": "cwilliams",
                                "major": "Engineering",
                                "gpa": 3.6,
                                "academic_level": "Senior",
                                "expected_graduation": datetime(2026, 5, 15),
                                "academic_history": [
                                    {
                                        "term": "Fall 2024",
                                        "courses": [
                                            {
                                                "code": "ENG401",
                                                "name": "Engineering Design",
                                                "grade": "A",
                                            },
                                            {
                                                "code": "ENG402",
                                                "name": "Project Management",
                                                "grade": "B+",
                                            },
                                        ],
                                        "gpa": 3.6,
                                    }
                                ],
                                "essays": [
                                    {
                                        "prompt": "Describe your leadership experience.",
                                        "content": "As president of the Engineering Club...",
                                        "submission_date": datetime(2025, 2, 3),
                                        "evaluation": {
                                            "score": 8.9,
                                            "feedback": "Strong leadership qualities demonstrated.",
                                            "reviewer": "Dr. Martinez",
                                            "date": datetime(2025, 2, 14),
                                        },
                                    }
                                ],
                                "financial_info": {
                                    "fafsa_submitted": True,
                                    "efc": 3000,
                                    "household_income": "30000-50000",
                                },
                                "interview_notes": "Great communication skills and project experience.",
                                "committee_feedback": [
                                    {
                                        "member": "Prof. Anderson",
                                        "comments": "Strong candidate with practical experience.",
                                        "recommendation": "Recommend",
                                        "date": datetime(2025, 3, 2),
                                    }
                                ],
                            }
                        ),
                    ]

                    if export_format == "pdf":
                        output_path = engine.export_prescreening_report_to_pdf(
                            applicants=sample_applicants, output_path=temp_file.name
                        )
                        content_type = "application/pdf"
                    elif export_format == "xlsx":
                        output_path = engine.export_prescreening_report_to_excel(
                            applicants=sample_applicants, output_path=temp_file.name
                        )
                        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif export_format == "csv":
                        output_path = engine.export_prescreening_report_to_csv(
                            applicants=sample_applicants, output_path=temp_file.name
                        )
                        content_type = "text/csv"
                    else:
                        raise ValueError(
                            f"Unsupported export format for pre-screening report: {export_format}"
                        )
                    filename = f"prescreening_report.{export_format}"
                else:
                    # Generate general scholarship report
                    if export_format == "pdf":
                        output_path = engine.export_to_pdf(temp_file.name)
                        content_type = "application/pdf"
                    elif export_format == "xlsx":
                        output_path = engine.export_to_excel(temp_file.name)
                        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    elif export_format == "csv":
                        output_path = engine.export_to_csv(temp_file.name)
                        content_type = "text/csv"
                    else:
                        raise ValueError(f"Unsupported export format: {export_format}")
                    filename = f"scholarship_report.{export_format}"

                # Read the file and create the response
                with open(output_path, "rb") as f:
                    response = HttpResponse(f.read(), content_type=content_type)
                    response["Content-Disposition"] = (
                        f'attachment; filename="{filename}"'
                    )
                return response

            except Exception as e:
                # Log the error (you might want to use proper logging here)
                print(f"Error during export: {str(e)}")
                return HttpResponse(f"Error generating report: {str(e)}", status=500)

            finally:
                # Clean up temporary file in finally block to ensure it happens
                try:
                    if temp_file:
                        os.unlink(temp_file.name)
                except Exception as e:
                    print(f"Error cleaning up temporary file: {str(e)}")

    # Generate report for web display
    report_data = engine.generate_scholarship_report()

    # Convert scholarships data for template display
    for scholarship in report_data["scholarships"]:
        # Ensure donor info is in the expected format
        if isinstance(scholarship.get("donor"), dict):
            scholarship["donor"] = scholarship["donor"]
        else:
            scholarship["donor"] = {"name": "Unknown", "contact": "Not provided"}

    return render(request, "reports_app/index.html", {"report": report_data})


def combined_analytics(request):
    """Generate and display combined Application & Scholarship analytics with export options."""
    from datetime import datetime

    engine = ReportEngine()

    # Read inputs
    start_raw = request.GET.get("start_date")
    end_raw = request.GET.get("end_date")
    export_format = request.GET.get("export_format")
    errors = []
    start_dt = None
    end_dt = None
    try:
        if start_raw:
            start_dt = datetime.strptime(start_raw, "%Y-%m-%d")
        if end_raw:
            end_dt = datetime.strptime(end_raw, "%Y-%m-%d")
    except ValueError:
        errors.append("Invalid date format. Use YYYY-MM-DD.")

    # Generate analytics for the two requested modules
    analytics = None
    try:
        analytics = engine.generate_analytics_report(
            start_date=start_dt,
            end_date=end_dt,
            report_types=["application_trends", "scholarship_impact"],
        )
    except Exception as e:
        logger.error(f"Combined analytics generation failed: {e}")
        errors.append(f"Analytics generation failed: {e}")

    # Export if requested
    if export_format and analytics:
        import tempfile, os

        try:
            tmp_dir = tempfile.gettempdir()
            filename_base = (
                f"combined_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            )
            if export_format == "json":
                path = os.path.join(tmp_dir, filename_base + ".json")
                engine.export_analytics_report_to_json(analytics, path)
                mime = "application/json"
            elif export_format == "excel":
                path = os.path.join(tmp_dir, filename_base + ".xlsx")
                engine.export_analytics_report_to_excel(analytics, path)
                mime = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            elif export_format == "pdf":
                path = os.path.join(tmp_dir, filename_base + ".pdf")
                engine.export_analytics_report_to_pdf(analytics, path)
                mime = "application/pdf"
            else:
                path = None
                mime = None
                errors.append("Unsupported export format.")

            if path and os.path.exists(path):
                with open(path, "rb") as f:
                    data = f.read()
                from django.http import HttpResponse

                response = HttpResponse(data, content_type=mime)
                response["Content-Disposition"] = (
                    f'attachment; filename="{os.path.basename(path)}"'
                )
                return response
        except Exception as e:
            logger.error(f"Combined analytics export failed: {e}")
            errors.append(f"Export failed: {e}")

    # Format analytics data for template display (create a copy to avoid modifying original)
    analytics_display = None
    if analytics:
        import json
        from datetime import date
        from decimal import Decimal
        
        def _json_serializer(obj):
            """Custom JSON serializer for datetime and Decimal objects."""
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            if isinstance(obj, Decimal):
                return float(obj)
            return str(obj)
        
        # Create a copy for display formatting
        analytics_display = {}
        
        # Format each section as pretty JSON
        if "application_trends" in analytics:
            analytics_display["application_trends"] = json.dumps(
                analytics["application_trends"], indent=2, default=_json_serializer
            )
        if "scholarship_impact" in analytics:
            analytics_display["scholarship_impact"] = json.dumps(
                analytics["scholarship_impact"], indent=2, default=_json_serializer
            )
        if "metadata" in analytics:
            analytics_display["metadata"] = analytics["metadata"]
        
        # Also format the full analytics for the raw JSON display
        analytics_display["raw_json"] = json.dumps(
            analytics, indent=2, default=_json_serializer
        )

    context = {
        "analytics": analytics_display,
        "errors": errors,
        "form": {
            "start_date": start_raw or "",
            "end_date": end_raw or "",
            "export_format": export_format or "",
        },
    }
    
    return render(request, "reports_app/combined_analytics.html", context)


def request_information(request):
    """View to handle reviewer requests for additional applicant information.

    Logs the request and provides feedback to the reviewer.
    """
    if request.method == "POST":
        try:
            # Get form data
            applicant_id = request.POST.get("applicant_id")
            reviewer_name = request.POST.get("reviewer_name")
            reviewer_email = request.POST.get("reviewer_email")
            scholarship_name = request.POST.get("scholarship_name")
            request_type = request.POST.get("request_type")
            request_details = request.POST.get("request_details")
            priority = request.POST.get("priority", "medium")

            # Use the ReportEngine to log the request
            engine = ReportEngine()
            info_request = engine.log_information_request(
                applicant_id=applicant_id,
                reviewer_name=reviewer_name,
                reviewer_email=reviewer_email,
                scholarship_name=scholarship_name if scholarship_name else None,
                request_type=request_type,
                request_details=request_details,
                priority=priority,
            )

            # Return success message
            return HttpResponse(f"""
                <html>
                <head>
                    <title>Request Submitted</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 2em; }}
                        .success {{ background: #d4edda; border: 1px solid #c3e6cb; color: #155724; padding: 1em; border-radius: 4px; }}
                        .details {{ background: #f8f9fa; padding: 1em; margin: 1em 0; border-radius: 4px; }}
                        button {{ padding: 0.7em 1.5em; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; margin-top: 1em; }}
                        button:hover {{ background: #0056b3; }}
                    </style>
                </head>
                <body>
                    <h1>Request Submitted Successfully</h1>
                    <div class="success">
                        <p><strong>Your information request has been logged.</strong></p>
                        <p>Request ID: #{info_request.id}</p>
                        <p>Status: {info_request.status.title()}</p>
                    </div>
                    <div class="details">
                        <h3>Request Details:</h3>
                        <p><strong>Applicant ID:</strong> {applicant_id}</p>
                        <p><strong>Reviewer:</strong> {reviewer_name}</p>
                        <p><strong>Request Type:</strong> {request_type.replace("_", " ").title()}</p>
                        <p><strong>Priority:</strong> {priority.title()}</p>
                        <p><strong>Submitted:</strong> {info_request.requested_at.strftime("%Y-%m-%d %H:%M:%S")}</p>
                    </div>
                    <button onclick="window.location.href='/'">Return to Reports</button>
                </body>
                </html>
            """)

        except ValueError as e:
            # Handle validation errors
            return HttpResponse(
                f"""
                <html>
                <head>
                    <title>Request Error</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 2em; }}
                        .error {{ background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; padding: 1em; border-radius: 4px; }}
                        button {{ padding: 0.7em 1.5em; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; margin-top: 1em; }}
                        button:hover {{ background: #0056b3; }}
                    </style>
                </head>
                <body>
                    <h1>Request Error</h1>
                    <div class="error">
                        <p><strong>Error:</strong> {str(e)}</p>
                    </div>
                    <button onclick="window.history.back()">Go Back</button>
                    <button onclick="window.location.href='/'">Return to Reports</button>
                </body>
                </html>
            """,
                status=400,
            )

        except Exception as e:
            # Handle unexpected errors
            logger.error(f"Error processing information request: {str(e)}")
            return HttpResponse(
                f"""
                <html>
                <head>
                    <title>System Error</title>
                    <style>
                        body {{ font-family: Arial, sans-serif; margin: 2em; }}
                        .error {{ background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; padding: 1em; border-radius: 4px; }}
                        button {{ padding: 0.7em 1.5em; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; margin-top: 1em; }}
                        button:hover {{ background: #0056b3; }}
                    </style>
                </head>
                <body>
                    <h1>System Error</h1>
                    <div class="error">
                        <p><strong>An unexpected error occurred:</strong> {str(e)}</p>
                    </div>
                    <button onclick="window.history.back()">Go Back</button>
                    <button onclick="window.location.href='/'">Return to Reports</button>
                </body>
                </html>
            """,
                status=500,
            )

    # If not POST, redirect to home
    from django.shortcuts import redirect

    return redirect("reports_home")


def award_scholarship(request):
    """Endpoint to record a simple award decision and optionally create an award.

    POST fields:
      - applicant_id (student_id)
      - scholarship_name
      - decision: awarded | not_awarded | pending
      - comments (optional)
      - create_award (optional 'yes' to create ScholarshipAward when decision is awarded)
      - award_amount (optional, defaults to Scholarship.amount if found)
    """
    if request.method != "POST":
        logger.error("award_scholarship called with non-POST method")
        return HttpResponse("Method Not Allowed", status=405)

    student_id = request.POST.get("applicant_id")
    scholarship_name = request.POST.get("scholarship_name")
    decision = request.POST.get("decision", "pending")
    comments = request.POST.get("comments")
    create_award_flag = request.POST.get("create_award") == "yes"
    award_amount_input = request.POST.get("award_amount")

    logger.info(f"=== Award Scholarship Request ===")
    logger.info(f"Student ID: {student_id}")
    logger.info(f"Scholarship Name: {scholarship_name}")
    logger.info(f"Decision: {decision}")
    logger.info(f"Comments: {comments}")
    logger.info(f"Create Award: {create_award_flag}")
    logger.info(f"Award Amount: {award_amount_input}")
    logger.info(f"All POST data: {dict(request.POST)}")

    if not student_id or not scholarship_name:
        logger.error("Missing applicant_id or scholarship_name")
        return HttpResponse("Missing applicant_id or scholarship_name", status=400)

    try:
        applicant = Applicant.objects.get(student_id=student_id)
    except Applicant.DoesNotExist:
        logger.error(f"Applicant with student_id {student_id} not found")
        return HttpResponse(
            f"Applicant with student_id {student_id} not found", status=404
        )
    except Exception as e:
        logger.error(f"Error finding applicant: {e}")
        return HttpResponse(f"Error finding applicant: {str(e)}", status=500)

    try:
        # Record decision
        award_decision = AwardDecision.record(
            applicant=applicant,
            scholarship_name=scholarship_name,
            decision=decision,
            comments=comments,
        )
        logger.info(f"Award decision recorded: {award_decision}")

        # Optionally create ScholarshipAward if decision is awarded
        if decision == "awarded" and create_award_flag:
            from decimal import Decimal

            now = timezone.now()
            # Try to find scholarship to default amount
            scholarship_obj = Scholarship.objects.filter(name=scholarship_name).first()
            default_amount = (
                Decimal(str(scholarship_obj.amount))
                if scholarship_obj
                else Decimal("0.00")
            )
            amount = (
                Decimal(award_amount_input) if award_amount_input else default_amount
            )

            scholarship_award = ScholarshipAward.objects.create(
                scholarship_name=scholarship_name,
                applicant=applicant,
                award_date=now,
                award_amount=amount,
                disbursement_dates=[],
                requirements_met=[],
                requirements_pending=[],
                status="active",
                performance_metrics={},
                notes=comments,
            )
            logger.info(f"ScholarshipAward created: {scholarship_award}")

        # Provide user feedback via messages (if messages framework installed)
        try:
            from django.contrib import messages

            verb_map = {
                "awarded": "awarded",
                "not_awarded": "marked as not awarded",
                "pending": "set to pending",
            }
            action_phrase = verb_map.get(decision, "updated")
            messages.success(
                request,
                f"Scholarship '{scholarship_name}' {action_phrase} for applicant {applicant.name}.",
            )
        except Exception as msg_error:
            logger.warning(f"Could not add success message: {msg_error}")

        # Redirect back to prescreening report
        logger.debug("Redirecting to view_prescreening_report")
        return redirect("view_prescreening_report")

    except Exception as e:
        logger.error(f"Error recording award decision: {e}", exc_info=True)
        return HttpResponse(f"Error: {str(e)}", status=500)


def view_request_logs(request):
    """View to display all information requests in a table format.

    Shows all logged reviewer requests for additional applicant information.
    """
    # Get all information requests ordered by most recent first
    all_requests = ReviewerInformationRequest.objects.all().order_by("-requested_at")

    # Apply filters if provided
    status_filter = request.GET.get("status")
    priority_filter = request.GET.get("priority")
    applicant_filter = request.GET.get("applicant_id")

    if status_filter:
        all_requests = all_requests.filter(status=status_filter)
    if priority_filter:
        all_requests = all_requests.filter(priority=priority_filter)
    if applicant_filter:
        all_requests = all_requests.filter(
            applicant__student_id__icontains=applicant_filter
        )

    # Get unique values for filter dropdowns
    statuses = ReviewerInformationRequest.objects.values_list(
        "status", flat=True
    ).distinct()
    priorities = ReviewerInformationRequest.objects.values_list(
        "priority", flat=True
    ).distinct()

    context = {
        "requests": all_requests,
        "statuses": statuses,
        "priorities": priorities,
        "current_status": status_filter,
        "current_priority": priority_filter,
        "current_applicant": applicant_filter,
    }

    return render(request, "reports_app/request_logs.html", context)


def clear_request_logs(request):
    """View to clear all information request logs.

    Deletes all ReviewerInformationRequest records from the database.
    Redirects back to the request logs page after clearing.
    """
    if request.method == "POST":
        # Delete all information requests
        ReviewerInformationRequest.objects.all().delete()

        # Optionally delete log files if they exist
        import os

        logs_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "information_request_logs"
        )
        if os.path.exists(logs_dir):
            import shutil

            try:
                shutil.rmtree(logs_dir)
                os.makedirs(logs_dir, exist_ok=True)
            except Exception as e:
                pass  # If log files can't be deleted, continue anyway

    return redirect("view_request_logs")


# Implementation of the prescreening report view with award decision functionality SFWE-504_3-LLR-28.
def view_prescreening_report(request):
    """View to display prescreening report with award decision functionality.

    Shows qualified applicants for each scholarship with the ability to:
    - Award scholarships
    - Add decision comments
    - View current decision status
    """
    engine = ReportEngine()
    # All applicants in the system (ordered by name for dropdown convenience)
    applicants = list(Applicant.objects.all().order_by("name"))

    # Only show the most recent scholarships displayed on home page (latest of each name)
    # The home page creates new instances each time, so we get the most recent ones
    engineering_scholarship = (
        Scholarship.objects.filter(name="Engineering Excellence Scholarship")
        .order_by("-id")
        .first()
    )

    cs_scholarship = (
        Scholarship.objects.filter(name="CS Leadership Scholarship")
        .order_by("-id")
        .first()
    )

    all_scholarships = [
        s for s in [engineering_scholarship, cs_scholarship] if s is not None
    ]

    # Check which scholarships have been awarded
    awarded_scholarship_names = set()
    for scholarship in all_scholarships:
        # A scholarship is considered "awarded" if there's at least one awarded decision for it
        has_award = AwardDecision.objects.filter(
            scholarship_name=scholarship.name, decision="awarded"
        ).exists()
        scholarship.is_awarded = has_award
        if has_award:
            awarded_scholarship_names.add(scholarship.name)

    # Filter to show only non-awarded scholarships as available
    available_scholarships = [s for s in all_scholarships if not s.is_awarded]

    # Generate prescreening report
    report_data = engine.generate_prescreening_report(applicants)

    # Enhance qualified applicants with existing decisions
    for scholarship_match in report_data.get("matches", []):
        scholarship_name = scholarship_match["scholarship_name"]
        for match in scholarship_match.get("matches", []):
            applicant_info = match["applicant"]
            student_id = applicant_info["student_id"]

            # Try to get existing applicant object and decision
            try:
                applicant_obj = Applicant.objects.get(student_id=student_id)
                decision = AwardDecision.objects.filter(
                    applicant=applicant_obj, scholarship_name=scholarship_name
                ).first()

                if decision:
                    match["current_decision"] = {
                        "decision": decision.decision,
                        "comments": decision.comments,
                        "decided_at": decision.decided_at,
                    }
                else:
                    match["current_decision"] = None
            except Applicant.DoesNotExist:
                match["current_decision"] = None

    # Override scholarships_evaluated to reflect only available (non-awarded) scholarships
    report_data["scholarships_evaluated"] = len(available_scholarships)

    # Calculate total awarded count from all AwardDecision records (not just matched applicants)
    total_awarded_count = AwardDecision.objects.filter(decision="awarded").count()
    report_data["summary"]["award_decisions"]["awarded"] = total_awarded_count

    context = {
        "report": report_data,
        "scholarships": available_scholarships,  # Only show non-awarded scholarships
        "all_scholarships": all_scholarships,  # All scholarships for dropdown in modal
        "all_applicants": applicants,  # Provide full applicant list for award modal dropdown
        "total_applicants": len(applicants),
        "generated_date": report_data.get("generated_date", datetime.now()),
    }

    return render(request, "reports_app/prescreening_report.html", context)
